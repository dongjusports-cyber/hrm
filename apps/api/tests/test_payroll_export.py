"""P5.2 — Xuất Excel bảng lương GenusSuite + audit log."""

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.modules.attendance.models import TimesheetMonth
from app.modules.attendance.timesheet import ensure_pay_period, rebuild_timesheets
from app.modules.core.export_log import ExportLog
from app.modules.mdm.models import Employee


def _hr_headers(client):
    token = client.post(
        "/api/auth/login", json={"username": "hr.demo", "password": "HrDemo@123456"}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _calc(client, db):
    ensure_pay_period(db, "2025-10")
    rebuild_timesheets(db, "2025-10", recalc_days=False)
    pay = ensure_pay_period(db, "2025-10")
    for emp in db.query(Employee).filter(Employee.deleted_at.is_(None)).all():
        ts = (
            db.query(TimesheetMonth)
            .filter(TimesheetMonth.pay_period_id == pay.id, TimesheetMonth.employee_id == emp.id)
            .one()
        )
        ts.worked_days = Decimal("26")
    # 1 NV CASH, còn lại ATM có TK
    cash = db.query(Employee).filter(Employee.employee_code == "1732").one()
    cash.pay_channel = "CASH"
    cash.bank_account = None
    atm = db.query(Employee).filter(Employee.employee_code == "5290").one()
    atm.pay_channel = "ATM"
    atm.bank_account = "0123456789"
    db.commit()

    from app.modules.attendance.timesheet import ensure_pay_period as ensure
    from app.modules.payroll import service as payroll_service

    def _noop(db_sess, period, *, recalc_days=True):
        ensure(db_sess, period)
        return type(
            "R",
            (),
            {"rows_upserted": 0, "message": "noop", "period": period, "pay_period_id": pay.id},
        )()

    original = payroll_service.rebuild_timesheets
    payroll_service.rebuild_timesheets = _noop
    try:
        assert (
            client.post(
                "/api/payroll/periods/2025-10/calculate", headers=_hr_headers(client)
            ).status_code
            == 200
        )
    finally:
        payroll_service.rebuild_timesheets = original


def test_export_atm_cash_and_audit(client, db):
    _calc(client, db)

    atm = client.get(
        "/api/payroll/periods/2025-10/export?channel=ATM",
        headers=_hr_headers(client),
    )
    assert atm.status_code == 200, atm.text
    assert "spreadsheetml" in atm.headers["content-type"]
    wb = load_workbook(BytesIO(atm.content))
    assert "ATM" in wb.sheetnames
    sheet = wb["ATM"]
    assert sheet.cell(row=7, column=1).value == "PAYROLL / BẢNG LƯƠNG"
    assert sheet.cell(row=1, column=1).value == "CÔNG TY TNHH DONGJU SPORTS VIỆT NAM"
    assert sheet.cell(row=10, column=1).value == "No"
    assert sheet.cell(row=12, column=3).value == "MSNV"
    assert "x1.5" in str(sheet.cell(row=12, column=29).value)
    assert "x2.1" in str(sheet.cell(row=12, column=31).value)
    assert "x2" in str(sheet.cell(row=12, column=33).value)
    assert "x3.5" in str(sheet.cell(row=12, column=35).value)
    assert "x5.1" in str(sheet.cell(row=12, column=43).value)
    assert sheet.cell(row=12, column=51).value == "Thực lãnh"
    assert sheet.cell(row=13, column=10).value == 10
    from app.modules.payroll.export_salary_table import (
        COL_WIDTHS,
        HEADER_EN_1,
        HEADER_EN_2,
        HEADER_ROW_VI,
        LAST_COL,
    )

    assert LAST_COL == 52
    assert len(HEADER_EN_1) == len(HEADER_EN_2) == len(HEADER_ROW_VI) == len(COL_WIDTHS) == LAST_COL
    hdr = sheet.cell(row=12, column=1)
    assert hdr.fill.start_color.rgb in ("00BDD7EE", "BDD7EE")
    company = sheet.cell(row=1, column=1)
    assert company.fill.start_color.rgb in ("000A4D8C", "0A4D8C")
    assert sheet.max_row >= 14

    cash = client.get(
        "/api/payroll/periods/2025-10/export?channel=CASH",
        headers=_hr_headers(client),
    )
    assert cash.status_code == 200
    wb_c = load_workbook(BytesIO(cash.content))
    assert "CASH" in wb_c.sheetnames
    codes = [
        wb_c["CASH"].cell(r, 3).value
        for r in range(12, wb_c["CASH"].max_row + 1)
        if wb_c["CASH"].cell(r, 3).value
    ]
    assert 1732 in codes or "1732" in codes

    both = client.get(
        "/api/payroll/periods/2025-10/export?channel=ALL",
        headers=_hr_headers(client),
    )
    assert both.status_code == 200
    wb_a = load_workbook(BytesIO(both.content))
    assert set(wb_a.sheetnames) >= {"TOTAL", "ATM", "CASH"}

    logs = db.query(ExportLog).filter(ExportLog.kind.like("payroll_%")).all()
    assert len(logs) >= 3
    assert all(l.period == "2025-10" for l in logs)


def test_export_bad_channel(client, db):
    res = client.get(
        "/api/payroll/periods/2025-10/export?channel=BTC",
        headers=_hr_headers(client),
    )
    assert res.status_code == 400


def test_export_filter_department_and_employee(client, db):
    _calc(client, db)
    from app.modules.mdm.models import Department

    dept = db.query(Department).filter(Department.code == "QC1").one()
    headers = _hr_headers(client)

    dept_res = client.get(
        f"/api/payroll/periods/2025-10/export?channel=ALL&department_id={dept.id}",
        headers=headers,
    )
    assert dept_res.status_code == 200, dept_res.text
    wb_d = load_workbook(BytesIO(dept_res.content))
    codes_d = {
        wb_d["TOTAL"].cell(r, 3).value
        for r in range(14, wb_d["TOTAL"].max_row)
        if wb_d["TOTAL"].cell(r, 3).value
    }
    assert codes_d == {5321} or codes_d == {"5321"}

    emp_res = client.get(
        "/api/payroll/periods/2025-10/export?channel=ALL&employee_code=5290",
        headers=headers,
    )
    assert emp_res.status_code == 200, emp_res.text
    wb_e = load_workbook(BytesIO(emp_res.content))
    codes_e = {
        wb_e["TOTAL"].cell(r, 3).value
        for r in range(14, wb_e["TOTAL"].max_row)
        if wb_e["TOTAL"].cell(r, 3).value
    }
    assert codes_e == {5290} or codes_e == {"5290"}
    assert "5290" in emp_res.headers.get("content-disposition", "")
    assert "10.2025" in emp_res.headers.get("content-disposition", "")
    assert "Dongju" in emp_res.headers.get("content-disposition", "")

    missing = client.get(
        "/api/payroll/periods/2025-10/export?channel=ALL&employee_code=999999",
        headers=headers,
    )
    assert missing.status_code == 404
