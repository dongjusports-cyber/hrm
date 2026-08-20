"""KPI giám đốc theo tổ — từ 8/2026, OT đủ loại, xuất Excel ngày/tháng."""

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.modules.attendance.models import AttendanceDay
from app.modules.mdm.models import Employee


def _hr_headers(client):
    token = client.post(
        "/api/auth/login", json={"username": "hr.demo", "password": "HrDemo@123456"}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _punch(db, code: str, day: date, *, ot_minutes: int = 0, punches: int = 2) -> None:
    emp = db.query(Employee).filter(Employee.employee_code == code).one()
    db.add(
        AttendanceDay(
            employee_id=emp.id,
            work_date=day,
            first_in=datetime(day.year, day.month, day.day, 1, 0, tzinfo=timezone.utc),
            last_out=datetime(day.year, day.month, day.day, 10, 0, tzinfo=timezone.utc),
            punch_count=punches,
            worked_hours=Decimal("8"),
            ot_minutes=ot_minutes,
            ot_on_books_minutes=ot_minutes,
            ot_external_minutes=0,
            is_workday=True,
        )
    )
    db.commit()


def test_kpi_day_rejects_before_august_2026(client):
    res = client.get("/api/reports/kpi/day?work_date=2026-07-31", headers=_hr_headers(client))
    assert res.status_code == 400, res.text
    assert "8/2026" in res.json()["detail"]


def test_kpi_day_and_month_by_team(client, db):
    day = date(2026, 8, 18)
    _punch(db, "5290", day, ot_minutes=180)
    _punch(db, "1514", day, ot_minutes=0)

    res = client.get("/api/reports/kpi/day?work_date=2026-08-18", headers=_hr_headers(client))
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["work_date"] == "2026-08-18"
    assert body["ot_people"] == 1
    assert float(body["ot_hours"]) == 3
    assert body["teams_with_ot"] >= 1
    sw = next(t for t in body["teams"] if t["department_code"] == "SW1")
    assert sw["present"] >= 2
    assert sw["ot_people"] == 1
    assert float(sw["ot_hours"]) == 3

    people = client.get(
        f"/api/reports/kpi/day/people?work_date=2026-08-18&team_id={sw['team_id']}",
        headers=_hr_headers(client),
    )
    assert people.status_code == 200
    codes = {p["employee_code"] for p in people.json()}
    assert "5290" in codes
    ot_row = next(p for p in people.json() if p["employee_code"] == "5290")
    assert float(ot_row["ot_hours"]) == 3

    month = client.get("/api/reports/kpi/month-teams?period=2026-08", headers=_hr_headers(client))
    assert month.status_code == 200, month.text
    m = month.json()
    assert m["period"] == "2026-08"
    assert float(m["ot_hours"]) == 3
    sw_m = next(t for t in m["teams"] if t["department_code"] == "SW1")
    cell = next(c for c in sw_m["days"] if c["work_date"] == "2026-08-18")
    assert cell["ot_people"] == 1
    assert float(cell["ot_hours"]) == 3
    assert len(sw_m["days"]) == 31


def test_kpi_export_day_and_month(client, db):
    _punch(db, "5290", date(2026, 8, 18), ot_minutes=120)
    headers = _hr_headers(client)

    day_x = client.get("/api/reports/kpi/export-day?work_date=2026-08-18", headers=headers)
    assert day_x.status_code == 200, day_x.text
    assert "spreadsheetml" in day_x.headers["content-type"]
    wb = load_workbook(BytesIO(day_x.content))
    assert "Theo to" in wb.sheetnames
    assert "Nguoi OT" in wb.sheetnames
    assert wb["Theo to"]["A1"].value and "DONGJU" in str(wb["Theo to"]["A1"].value).upper()

    month_x = client.get("/api/reports/kpi/export-month?period=2026-08", headers=headers)
    assert month_x.status_code == 200, month_x.text
    wb2 = load_workbook(BytesIO(month_x.content))
    assert set(wb2.sheetnames) >= {"Chuyen can", "Tang ca", "Nghi viec", "Tong hop"}
    # Đủ cột ngày trong tháng
    headers_row = [c.value for c in wb2["Chuyen can"][5]]
    assert "01" in headers_row and "31" in headers_row
