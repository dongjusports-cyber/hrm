"""OT ngoài — làm tròn 30p + công thức 22§22.8 (không qua payslip)."""

from io import BytesIO
from decimal import Decimal

from openpyxl import load_workbook

from app.modules.payroll.engine_allowances import AllowanceLine
from app.modules.payroll.engine_ot import OtHours, OtInput, compute_ot_pay, quantize_ot_hours
from app.modules.payroll.ot_external import (
    OtExternalPayRow,
    OtExternalSummary,
    _external_rate,
    build_ot_external_excel,
    split_external_ot_hours,
)
from app.modules.policy.seed_payload import default_payload


def test_external_rate_from_policy():
    policy = default_payload()
    assert _external_rate(policy) == Decimal("1.5")
    policy["ot_split"]["ot_external"]["rate_key"] = "weekday"
    assert _external_rate(policy) == Decimal("1.5")


def test_external_hours_quantized_per_minute():
    policy = default_payload()
    assert quantize_ot_hours(Decimal("2.4"), policy) == Decimal("2.40")
    assert quantize_ot_hours(Decimal("0.25"), policy) == Decimal("0.25")


def test_external_pay_same_formula_as_weekday_ot():
    lines = [
        AllowanceLine("TOXIC", "Độc hại", Decimal("100000"), Decimal("100000"), True, True),
    ]
    policy = default_payload()
    hours = Decimal("2.0")
    r = compute_ot_pay(
        OtInput(
            contract_salary=Decimal("5675000"),
            salary_divisor=Decimal("26"),
            allowance_lines=lines,
            attend_full_monthly=Decimal("0"),
            hours=OtHours(weekday=hours),
            policy=policy,
        )
    )
    assert r.ot_pay > 0
    assert r.detail["effective_hours"]["weekday"] == "2.00"


def test_split_external_hours_new_rebuild_includes_sunday():
    h = split_external_ot_hours(
        external=Decimal("11.17"),
        weekend=Decimal("9.17"),
        holiday=Decimal("0"),
    )
    assert h.weekday == Decimal("2.00")
    assert h.weekend == Decimal("9.17")
    assert h.holiday == Decimal("0")


def test_split_external_hours_legacy_weekday_plus_sunday():
    """Dữ liệu cũ: external chỉ ngày thường, CN nằm weekend — không trừ âm."""
    h = split_external_ot_hours(
        external=Decimal("2"),
        weekend=Decimal("4"),
        holiday=Decimal("0"),
    )
    assert h.weekday == Decimal("2")
    assert h.weekend == Decimal("4")


def test_split_external_hours_sunday_only_legacy():
    h = split_external_ot_hours(
        external=Decimal("0"),
        weekend=Decimal("9.17"),
        holiday=Decimal("0"),
    )
    assert h.weekday == Decimal("0")
    assert h.weekend == Decimal("9.17")


def _cell_fill(cell) -> str:
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return str(rgb or "")[-6:].upper()


def test_ot_external_excel_matches_print_layout():
    summary = OtExternalSummary(
        period="2026-08",
        employee_count=1,
        total_raw_hours=Decimal("2.00"),
        total_effective_hours=Decimal("2.00"),
        total_amount_vnd=Decimal("150000"),
        rows=[
            OtExternalPayRow(
                employee_code="5290",
                full_name="Nguyen Van A",
                bank_account="123456",
                raw_hours=Decimal("2.00"),
                effective_hours=Decimal("2.00"),
                ot_base=Decimal("5675000"),
                hourly_base=Decimal("25000"),
                rate=Decimal("1.5"),
                amount_vnd=Decimal("150000"),
                hours_x15=Decimal("2.00"),
                pay_x15=Decimal("150000"),
            )
        ],
        policy_note="OT ngoài: x1,5 · x2 · x2,1 · x3.",
    )
    wb = load_workbook(BytesIO(build_ot_external_excel(summary)))
    ws = wb.active
    assert ws["A1"].value and "DONGJU" in str(ws["A1"].value).upper()
    assert _cell_fill(ws["A1"]) == "0A4D8C"
    assert ws["A1"].font.color.rgb[-6:].upper() == "FFFFFF"
    assert ws["A5"].value and "OT NGOÀI" in str(ws["A5"].value)
    assert "THÁNG 08" in str(ws["A6"].value)
    assert ws["A10"].value == "STT"
    assert "x1.5" in str(ws["H10"].value)
    assert "x2" in str(ws["J10"].value)
    assert "x2.1" in str(ws["L10"].value)
    assert "x3" in str(ws["N10"].value)
    assert ws["P10"].value and "Tổng tiền" in str(ws["P10"].value)
    assert _cell_fill(ws["A10"]) == "BDD7EE"
    assert _cell_fill(ws["P10"]) == "BDD7EE"
    assert ws["B12"].value == "5290"
    assert ws["H12"].value == 2.0
    assert ws["I12"].value == 150000
    assert ws["P12"].value == 150000
    assert ws.freeze_panes == "A12"
    footer = 13
    assert "Tổng cộng" in str(ws.cell(row=footer, column=1).value)
    assert _cell_fill(ws.cell(row=footer, column=1)) == "5B9BD5"


def test_ot_external_excel_mixed_rates():
    summary = OtExternalSummary(
        period="2026-08",
        employee_count=1,
        total_raw_hours=Decimal("12.00"),
        total_effective_hours=Decimal("12.00"),
        total_amount_vnd=Decimal("900000"),
        rows=[
            OtExternalPayRow(
                employee_code="5290",
                full_name="Nguyen Van A",
                bank_account="123456",
                raw_hours=Decimal("12.00"),
                effective_hours=Decimal("12.00"),
                ot_base=Decimal("5675000"),
                hourly_base=Decimal("25000"),
                rate=Decimal("2.0"),
                amount_vnd=Decimal("900000"),
                hours_x15=Decimal("2.00"),
                pay_x15=Decimal("75000"),
                hours_x20=Decimal("8.00"),
                pay_x20=Decimal("400000"),
                hours_x21=Decimal("0"),
                pay_x21=Decimal("0"),
                hours_x30=Decimal("2.00"),
                pay_x30=Decimal("150000"),
            )
        ],
        policy_note="chi tiết hệ số",
    )
    ws = load_workbook(BytesIO(build_ot_external_excel(summary))).active
    assert ws["H12"].value == 2.0
    assert ws["J12"].value == 8.0
    assert ws["N12"].value == 2.0
    assert ws["P12"].value == 900000
    assert ws.cell(row=13, column=8).value == 2.0
    assert ws.cell(row=13, column=10).value == 8.0
    assert ws.cell(row=13, column=16).value == 900000

