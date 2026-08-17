"""Parse Thông tin bổ sung — hôn nhân Excel True=-1, STK không lấy nhầm SĐT."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook

from app.scripts.thong_tin_bo_sung import (
    DEFAULT_ETHNICITY,
    DEFAULT_NATIONALITY,
    DEFAULT_RELIGION,
    apply_supplement_to_mapping,
    load_supplement,
    resolve_gender_flag,
    resolve_marital_status,
    split_pccc_hse,
    _bank_account,
    _prefer_longer_id,
)


def test_marital_excel_true_is_minus_one():
    assert resolve_marital_status(-1) == "married"
    assert resolve_marital_status(1) == "married"
    assert resolve_marital_status(0) == "single"
    assert resolve_marital_status(None) is None


def test_gender_flag_0_female_1_male():
    assert resolve_gender_flag(0) == "F"
    assert resolve_gender_flag(1) == "M"
    assert resolve_gender_flag("Female") == "F"


def test_bank_skips_when_same_as_phone():
    assert _bank_account("0338434799", "0338434799") is None
    assert _bank_account("060227267677", "0908275468") == "060227267677"


def test_si_book_keeps_longer_prefix():
    assert _prefer_longer_id("72120192380", "7212019238") == "72120192380"
    assert _prefer_longer_id(None, "7212019238") == "7212019238"
    assert _prefer_longer_id("7212019238", "72120192380") == "72120192380"


def test_apply_overwrites_marital_fills_phone():
    dest = {
        "gender": "F",
        "marital_status": None,
        "children_count": 0,
        "phone": None,
        "bank_account": "060227267677",
        "si_book_no": "72120192380",
    }
    rec = {
        "gender": "F",
        "marital_status": "married",
        "children_count": 2,
        "phone": "0908275468",
        "bank_account": "060227267677",
        "si_book_no": "7212019238",
    }
    changed = apply_supplement_to_mapping(dest, rec)
    assert dest["marital_status"] == "married"
    assert dest["children_count"] == 2
    assert dest["phone"] == "0908275468"
    assert dest["si_book_no"] == "72120192380"
    assert dest["bank_account"] == "060227267677"
    assert "marital_status" in changed
    assert "si_book_no" not in changed


def test_load_supplement_from_fixture_xlsx(tmp_path: Path):
    path = tmp_path / "bo_sung.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "giới tính- hôn nhân"
    ws.append([None, None, "0 = nữ / 1 = nam", "0 = độc thân / 1 = đã kết hôn"])
    ws.append([None, None, "giới tính", "hôn nhân", "số con"])
    ws.append([1496, "LÊ KIỀU NHƯ", 0, -1, 2])
    ws.append([1514, "NGUYỄN THỊ NGỌC LUYẾN", 0, 0, None])
    ws2 = wb.create_sheet("02")
    ws2.append(["THE LIST"])
    ws2.append([])
    ws2.append(
        [
            "Dept",
            "EmpID",
            "Full Name",
            "Ngân hàng",
            "Birthday",
            "Join Date",
            "Place Of Birth",
            "Basic Salary",
            "Group",
            "Sex",
            "Position",
            "Tel",
            "Education",
            "Addr",
            "PersonID",
            "Place_per_ID",
            "Issue_DT",
            "Social No",
            "Contract No",
            "Start Contract",
            "End Contract",
            "Left Date",
        ]
    )
    ws2.append(
        [
            "HR & Admin",
            1496,
            "LÊ KIỀU NHƯ",
            "060227267677",
            "21/01/1989",
            "27/02/2015",
            "Tây Ninh",
            7885000,
            "HR/Admin Staff",
            "Female",
            "Staff",
            908275468,
            "University",
            "GIA TÂN",
            "072189002248",
            "Tây Ninh",
            date(2022, 3, 1),
            72120192380,
            "1496/VTH",
            date(2017, 3, 27),
            None,
            None,
        ]
    )
    ws3 = wb.create_sheet("03")
    ws3.append(["BASIC SALARY SUMMARY REPORT"])
    ws3.append([])
    ws3.append(
        [
            "NO",
            "DEPARTMENT",
            "GROUP",
            "EMP ID",
            "FULL NAME",
            "JOIN DATE",
            "POSITION",
            "PRO SALARY",
            "BASIC SALARY",
            "PCCC+HSE_AMT",
            "POS_AMT",
            "TOXIC_AMT",
            "INDUS_AMT",
            "TRANS_AMT",
            "TECH_AMT",
            "OTHER_AMT",
            "TOTAL",
            "UNION",
        ]
    )
    ws3.append(
        [
            1,
            "HR & Admin",
            "HR/Admin Staff",
            1496,
            "LÊ KIỀU NHƯ",
            "27/02/2015",
            "Staff",
            4080000,
            7885000,
            50000,
            0,
            0,
            600000,
            800000,
            500000,
            300000,
            0,
            "N",
        ]
    )
    wb.save(path)
    wb.close()

    recs = load_supplement(path)
    assert recs["1496"]["marital_status"] == "married"
    assert recs["1496"]["children_count"] == 2
    assert recs["1496"]["gender"] == "F"
    assert recs["1496"]["phone"] == "0908275468"
    assert recs["1496"]["bank_account"] == "060227267677"
    assert recs["1496"]["education_code"] == "EDUCATION_LEVEL008"
    assert recs["1496"]["nationality_code"] == DEFAULT_NATIONALITY
    assert recs["1496"]["ethnicity_code"] == DEFAULT_ETHNICITY
    assert recs["1496"]["religion_code"] == DEFAULT_RELIGION
    codes = {a["component_code"]: a["amount"] for a in recs["1496"]["allowances"]}
    assert codes["HSE"] == "50000.00"
    assert codes["ATTEND"] == "600000.00"
    assert codes["TRANSPORT"] == "800000.00"
    assert codes["TECH"] == "500000.00"
    assert codes["OTHER"] == "300000.00"
    assert "PCCC" not in codes
    assert recs["1514"]["marital_status"] == "single"
    assert recs["1514"]["children_count"] == 0


def test_split_pccc_hse():
    from decimal import Decimal

    assert split_pccc_hse(Decimal("50000")) == {"HSE": Decimal("50000")}
    assert split_pccc_hse(Decimal("882000")) == {"PCCC": Decimal("882000")}
    assert split_pccc_hse(Decimal("932000")) == {
        "HSE": Decimal("50000"),
        "PCCC": Decimal("882000"),
    }


def test_education_overwrite_and_identity_defaults():
    dest = {"education_code": "EDUCATION_LEVEL003", "nationality_code": None}
    rec = {"education_code": "EDUCATION_LEVEL008"}
    apply_supplement_to_mapping(dest, rec)
    assert dest["education_code"] == "EDUCATION_LEVEL008"
    assert dest["nationality_code"] == DEFAULT_NATIONALITY
    assert dest["ethnicity_code"] == DEFAULT_ETHNICITY
    assert dest["religion_code"] == DEFAULT_RELIGION
