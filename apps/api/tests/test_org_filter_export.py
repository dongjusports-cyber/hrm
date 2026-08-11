"""Hạng mục 1.4 — API + lưới danh sách NV: bộ lọc Bộ phận › Tổ, xuất Excel theo bộ lọc."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.modules.mdm.models import Department, Employee, Team


def _hr_headers(client):
    token = client.post(
        "/api/auth/login", json={"username": "hr.demo", "password": "HrDemo@123456"}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_team(db: Session, dept_code: str, team_code: str, team_name: str) -> Team:
    dept = db.query(Department).filter(Department.code == dept_code).one()
    team = Team(department_id=dept.id, code=team_code, name=team_name)
    db.add(team)
    db.flush()
    db.commit()
    return team


def test_list_teams(client, db: Session):
    _make_team(db, "SW1", "SW1-T1", "Sewing Team 1")
    headers = _hr_headers(client)
    res = client.get("/api/teams", headers=headers)
    assert res.status_code == 200, res.text
    codes = {t["code"] for t in res.json()}
    assert "SW1-T1" in codes
    row = next(t for t in res.json() if t["code"] == "SW1-T1")
    assert row["department_code"] == "SW1"


def test_filter_by_department(client, db: Session):
    headers = _hr_headers(client)
    dept = db.query(Department).filter(Department.code == "QC1").one()
    res = client.get(f"/api/employees?department_id={dept.id}", headers=headers)
    assert res.status_code == 200, res.text
    codes = {e["employee_code"] for e in res.json()}
    assert codes == {"5321"}


def test_filter_by_team(client, db: Session):
    team = _make_team(db, "SW1", "SW1-T2", "Sewing Team 2")
    emp = db.query(Employee).filter(Employee.employee_code == "1514").one()
    emp.team_id = team.id
    db.commit()

    headers = _hr_headers(client)
    res = client.get(f"/api/employees?team_id={team.id}", headers=headers)
    assert res.status_code == 200, res.text
    codes = {e["employee_code"] for e in res.json()}
    assert codes == {"1514"}
    row = next(e for e in res.json() if e["employee_code"] == "1514")
    assert row["team_code"] == "SW1-T2"
    assert row["team_name"] == "Sewing Team 2"


def test_seniority_and_contract_type_labels(client, db: Session):
    emp = db.query(Employee).filter(Employee.employee_code == "1514").one()
    emp.join_date = date(2020, 1, 15)
    emp.status = "active"
    db.commit()

    headers = _hr_headers(client)
    res = client.get("/api/employees?q=1514", headers=headers)
    row = next(e for e in res.json() if e["employee_code"] == "1514")
    assert row["contract_type_label"] == "Chính thức"
    assert "năm" in row["seniority_label"]


def test_export_respects_filter_and_columns(client, db: Session):
    dept = db.query(Department).filter(Department.code == "SW1").one()
    headers = _hr_headers(client)
    res = client.get(
        f"/api/employees/export?department_id={dept.id}&columns=employee_code,full_name,contract_salary",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["MSNV", "Họ tên", "Lương HĐ"]

    sw1_codes = {
        e.employee_code
        for e in db.query(Employee).filter(Employee.department_id == dept.id).all()
    }
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == len(sw1_codes)
    exported_codes = {r[0] for r in data_rows}
    assert exported_codes == sw1_codes


def test_export_default_columns_when_missing(client, db: Session):
    headers = _hr_headers(client)
    res = client.get("/api/employees/export", headers=headers)
    assert res.status_code == 200, res.text
    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "MSNV"
    assert "Tài khoản" in header
