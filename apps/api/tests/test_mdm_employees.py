"""P1.4 — Employees + Departments + import Excel."""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook


def _hr_headers(client):
    token = client.post(
        "/api/auth/login", json={"username": "hr.demo", "password": "HrDemo@123456"}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def test_seed_fixture_employees(client):
    res = client.get("/api/employees", headers=_hr_headers(client))
    assert res.status_code == 200
    codes = {e["employee_code"] for e in res.json()}
    assert {"1514", "1643", "5290", "5321", "1732"} <= codes
    emp = next(e for e in res.json() if e["employee_code"] == "5290")
    assert Decimal(str(emp["contract_salary"])) == Decimal("5675000")


def test_create_and_update_employee(client):
    headers = _hr_headers(client)
    created = client.post(
        "/api/employees",
        headers=headers,
        json={
            "employee_code": "9001",
            "full_name": "Test NV",
            "team_code": "T1",
            "department_code": "SW1",
            "contract_salary": "6000000",
            "probation_salary": "5000000",
            "pay_channel": "CASH",
        },
    )
    assert created.status_code == 201, created.text
    assert created.json()["pay_channel"] == "CASH"
    assert created.json()["team_code"] == "T1"
    assert created.json()["department_code"] == "SW1"
    assert Decimal(str(created.json()["contract_salary"])) == Decimal("6000000")

    uid = created.json()["id"]
    updated = client.put(
        f"/api/employees/{uid}",
        headers=headers,
        json={"full_name": "Test NV 2", "contract_salary": "6100000"},
    )
    assert updated.status_code == 200
    assert updated.json()["full_name"] == "Test NV 2"


def test_create_employee_with_allowances(client):
    headers = _hr_headers(client)
    created = client.post(
        "/api/employees",
        headers=headers,
        json={
            "employee_code": "9003",
            "full_name": "NV Phu Cap",
            "team_code": "T1",
            "department_code": "SW1",
            "contract_salary": "6000000",
            "allowances": [
                {"allowance_code": "ATTEND", "amount": "600000"},
                {"allowance_code": "TRANSPORT", "amount": "800000"},
            ],
        },
    )
    assert created.status_code == 201, created.text
    listed = client.get("/api/payroll/allowances?employee_code=9003", headers=headers)
    assert listed.status_code == 200, listed.text
    rows = listed.json()
    by_code = {r["allowance_code"]: Decimal(str(r["amount"])) for r in rows}
    assert by_code["ATTEND"] == Decimal("600000")
    assert by_code["TRANSPORT"] == Decimal("800000")


def test_create_employee_requires_team(client):
    """23§: mọi NV thuộc về một tổ — không cho tạo NV thiếu team (21§21.3, dọn dẹp đợt 1)."""
    headers = _hr_headers(client)
    res = client.post(
        "/api/employees",
        headers=headers,
        json={"employee_code": "9002", "full_name": "Thiếu Tổ", "contract_salary": "6000000"},
    )
    assert res.status_code == 400
    assert "tổ" in res.json()["detail"].lower()


def test_import_excel(client):
    headers = _hr_headers(client)
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "employee_code",
            "full_name",
            "team_code",
            "department_code",
            "contract_salary",
            "probation_salary",
            "pay_channel",
        ]
    )
    ws.append(["8801", "Import One", "T1", "SW1", "7000000", "6000000", "ATM"])
    ws.append(["5290", "Lê Văn C Updated", "T1", "SW1", "5675000", "4840750", "ATM"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    res = client.post(
        "/api/employees/import",
        headers=headers,
        files={
            "file": (
                "nv.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["created"] >= 1
    assert body["updated"] >= 1

    listed = client.get("/api/employees?q=8801", headers=headers).json()
    row = next(e for e in listed if e["employee_code"] == "8801")
    assert row["team_code"] == "T1"
    assert row["department_code"] == "SW1"


def test_search_employee(client):
    headers = _hr_headers(client)
    res = client.get("/api/employees?q=5290", headers=headers)
    assert res.status_code == 200
    assert len(res.json()) >= 1
    assert res.json()[0]["employee_code"] == "5290"


def test_hr_cannot_mutate_departments(client):
    headers = _hr_headers(client)
    bad = client.post(
        "/api/departments",
        headers=headers,
        json={"code": "X99", "name": "Forbidden", "category": "direct"},
    )
    assert bad.status_code == 403


def test_admin_department_crud(client):
    token = client.post(
        "/api/auth/login", json={"username": "admin", "password": "Admin@DongJu2026"}
    ).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    created = client.post(
        "/api/departments",
        headers=headers,
        json={"code": "ADM1", "name": "Admin Dept", "category": "admin_indirect"},
    )
    assert created.status_code == 201, created.text
    dept_id = created.json()["id"]
    updated = client.put(
        f"/api/departments/{dept_id}",
        headers=headers,
        json={"name": "Admin Dept 2"},
    )
    assert updated.status_code == 200
    assert updated.json()["name"] == "Admin Dept 2"
    deleted = client.delete(f"/api/departments/{dept_id}", headers=headers)
    assert deleted.status_code == 200


def test_import_resign_date(client):
    headers = _hr_headers(client)
    wb = Workbook()
    ws = wb.active
    ws.append(
        ["employee_code", "full_name", "team_code", "department_code", "contract_salary", "resign_date"]
    )
    ws.append(["8802", "Nghi Viec", "T1", "SW1", "5000000", "2025-10-15"])
    buf = BytesIO()
    wb.save(buf)
    res = client.post(
        "/api/employees/import",
        headers=headers,
        files={
            "file": (
                "resign.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert res.status_code == 200, res.text
    listed = client.get("/api/employees?q=8802", headers=headers).json()
    emp = next(e for e in listed if e["employee_code"] == "8802")
    assert emp["resign_date"] == "2025-10-15"
    assert emp["status"] == "resigned"


def test_employee_import_template_xlsx(client):
    headers = _hr_headers(client)
    res = client.get("/api/employees/import-template", headers=headers)
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers.get("content-type", "")
    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    headers_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "MSNV" in headers_row
    assert "Họ tên" in headers_row
    assert "Tình trạng hôn nhân" in headers_row
    assert "Địa chỉ tạm trú" in headers_row
    assert "Huong dan" in wb.sheetnames


def test_import_excel_full_profile_vietnamese(client):
    headers = _hr_headers(client)
    from app.modules.mdm.import_excel import build_employee_import_template

    content = build_employee_import_template()
    res = client.post(
        "/api/employees/import",
        headers=headers,
        files={
            "file": (
                "mau.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["created"] >= 1
    listed = client.get("/api/employees?q=8810", headers=headers).json()
    emp = next(e for e in listed if e["employee_code"] == "8810")
    assert emp["full_name"] == "NGUYỄN VĂN MẪU"
    assert emp["gender"] == "male"
    assert emp["phone"] == "0901234567"
    assert emp["team_code"] == "T1"
    assert Decimal(str(emp["contract_salary"])) == Decimal("6500000")
    assert emp["si_enrolled"] is True
    assert emp["marital_status"] == "single"
    assert emp["nationality_code"] == "NATIONALITY001"
    assert emp["ethnicity_code"] == "ETHNICITY001"
    assert emp["religion_code"] == "RELIGION001"
    assert emp["education_code"] == "EDUCATION_LEVEL004"
    assert emp["birth_place_code"] == "BIRTH_PLACE030"
    assert emp["id_issue_place_code"] == "ID_ISSUE_PLACE035"
    assert emp["id_number"] == "079098001234"
    assert emp["permanent_address"]
    assert emp["status"] == "probation"

