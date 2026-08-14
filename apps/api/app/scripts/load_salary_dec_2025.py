"""
[DEPRECATED — V1, đã thay bằng app/scripts/import_genussuite_2026.py ở hạng mục 1.3]
Script V1 này tự tạo Department giả từ tên cột trong file lương, KHÔNG gán Tổ (team_id) —
không còn khớp thiết kế V2 (21§21.3: NV thuộc Tổ, bộ phận suy ra qua Tổ). Giữ lại để tham
khảo lịch sử, KHÔNG chạy lại trên dữ liệu V2 hiện tại (sẽ tạo department rác không có tổ).

Nạp dữ liệu test từ bảng lương Dec.2025
→ departments + employees + phụ cấp gán + timesheet kỳ 2025-12.

Nguồn chuẩn trên máy Chủ (.123):
  C:\\DATA\\HRM\\dj-hrm\\dj-hrm\\HIEN_PHAP\\Salary\\  (hoặc file ngoài repo tại D:\\HRM\\)

Chạy (Docker, máy dev .123):
  docker cp "C:/DATA/HRM/dj-hrm/dj-hrm/HIEN_PHAP/Salary/2.Salary table for Dec.2025.xlsx" djhrm-api:/tmp/salary_dec.xlsx
  docker compose exec api python -m app.scripts.load_salary_dec_2025 /tmp/salary_dec.xlsx
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.modules.attendance.models import TimesheetMonth
from app.modules.attendance.timesheet import ensure_pay_period, seed_leave_types
from app.modules.mdm.models import Department, Employee
from app.modules.mdm.service import get_or_create_department_by_code
from app.modules.payroll.models import AllowanceType, EmployeeAllowanceAssignment
from app.modules.payroll.money import D, ZERO, money_vnd
from app.modules.payroll.seed_allowances import seed_allowance_types
from app.modules.policy.service import seed_default_package

PERIOD = "2025-12"
DIVISOR = Decimal("26")  # rule 27→26 tháng 12/2025 (OFF thường 27)
DATA_START_ROW = 14


def _slug_dept(section: str) -> str:
    raw = re.sub(r"[^A-Za-z0-9]+", "_", (section or "UNK").strip()).strip("_").upper()
    return (raw[:24] or "UNK")


def _dept_category(section: str) -> str:
    s = (section or "").lower()
    if any(k in s for k in ("account", "hr", "admin", "it", "office", "management")):
        return "admin_indirect"
    if any(k in s for k in ("qc", "kcs", "warehouse", "store", "wh", "cutting")):
        return "prod_indirect"
    return "direct"


def _parse_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _money(val) -> Decimal:
    if val is None or val == "":
        return ZERO
    try:
        return money_vnd(D(val))
    except Exception:
        return ZERO


def _dec(val) -> Decimal:
    if val is None or val == "":
        return ZERO
    try:
        return D(val)
    except Exception:
        return ZERO


def _code(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, float):
        if val != int(val):
            return str(val).strip()
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    text = str(val).strip()
    if not text or text.lower() in ("msnv", "staff no", "code"):
        return None
    try:
        return str(int(float(text)))
    except ValueError:
        return text


def _gender(val) -> str | None:
    t = str(val or "").strip().lower()
    if not t:
        return None
    if "nam" in t and "nữ" not in t and "nu" not in t:
        return "M"
    if "nữ" in t or "nu" in t or "female" in t:
        return "F"
    return None


def _monthly_from_prorata(excel_amt: Decimal, worked: Decimal) -> Decimal:
    """Excel tháng → mức tháng đầy đủ (ước lượng /26 × công)."""
    if excel_amt <= 0:
        return ZERO
    if worked <= 0:
        return money_vnd(excel_amt)
    return money_vnd(excel_amt * DIVISOR / worked)


def _upsert_allowance(
    db: Session,
    emp: Employee,
    types: dict[str, AllowanceType],
    code: str,
    amount: Decimal,
) -> str:
    at = types.get(code)
    if at is None or amount < 0:
        return "skip"
    row = (
        db.query(EmployeeAllowanceAssignment)
        .filter(
            EmployeeAllowanceAssignment.employee_id == emp.id,
            EmployeeAllowanceAssignment.allowance_type_id == at.id,
        )
        .one_or_none()
    )
    if row is None:
        db.add(
            EmployeeAllowanceAssignment(
                employee_id=emp.id,
                allowance_type_id=at.id,
                amount=amount,
            )
        )
        return "created"
    row.amount = amount
    return "updated"


def load(db: Session, path: Path) -> dict:
    seed_default_package(db)
    seed_leave_types(db)
    seed_allowance_types(db)
    types = {t.code: t for t in db.query(AllowanceType).all()}
    pay = ensure_pay_period(db, PERIOD)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    emp_created = emp_updated = 0
    asg_created = asg_updated = 0
    ts_upserted = 0
    skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DATA_START_ROW or not row:
            continue
        code = _code(row[2] if len(row) > 2 else None)
        name = str(row[3] if len(row) > 3 else "").strip()
        if not code or not name:
            skipped += 1
            continue

        section = str(row[1] or "UNK").strip()
        dept_code = _slug_dept(section)
        dept = get_or_create_department_by_code(db, dept_code, name=section)
        dept.category = _dept_category(section)
        if section and section not in (dept.mitapro_names or []):
            names = list(dept.mitapro_names or [])
            names.append(section)
            dept.mitapro_names = names

        bank = str(row[5] or "").strip() if len(row) > 5 else ""
        gender = _gender(row[36] if len(row) > 36 else None)
        values = {
            "full_name": name,
            "gender": gender,
            "id_number": (str(row[4]).strip() if len(row) > 4 and row[4] else None),
            "bank_account": bank or None,
            "pay_channel": "ATM" if bank else "CASH",
            # V1 cũ gán department_id trực tiếp — cột đã xóa ở V2 (21§21.3). Script này
            # DEPRECATED, không dùng nữa nên không gán team_id thay thế (xem docstring đầu file).
            "position_title": (str(row[7]).strip() if len(row) > 7 and row[7] else None),
            "join_date": _parse_date(row[8] if len(row) > 8 else None),
            "contract_signed_at": _parse_date(row[9] if len(row) > 9 else None),
            "probation_salary": _money(row[10] if len(row) > 10 else None),
            "contract_salary": _money(row[11] if len(row) > 11 else None),
            "status": "active",
            "si_enrolled": True,
        }

        emp = (
            db.query(Employee)
            .filter(Employee.employee_code == code, Employee.deleted_at.is_(None))
            .one_or_none()
        )
        if emp is None:
            emp = Employee(employee_code=code, **values)
            db.add(emp)
            db.flush()
            emp_created += 1
        else:
            for k, v in values.items():
                setattr(emp, k, v)
            emp.deleted_at = None
            emp_updated += 1

        # Cột Excel (1-based): M=13 Pro, N=14 OFF, O=15 Total ngày công ← worked_days
        #                     P=16 AL, Q=17 REM, AB=28 OT giờ thường
        total_d = _dec(row[14] if len(row) > 14 else None)  # cột O
        al = _dec(row[15] if len(row) > 15 else None)  # cột P
        rem = _dec(row[16] if len(row) > 16 else None)  # cột Q
        attend = _money(row[18] if len(row) > 18 else None)
        position = _money(row[19] if len(row) > 19 else None)
        toxic = _money(row[20] if len(row) > 20 else None)
        trans = _money(row[21] if len(row) > 21 else None)
        pccc = _money(row[22] if len(row) > 22 else None)
        tech = _money(row[23] if len(row) > 23 else None)
        senior = _money(row[24] if len(row) > 24 else None)
        other = _money(row[25] if len(row) > 25 else None)
        ot_h = _dec(row[27] if len(row) > 27 else None)

        # Mức tháng: ATTEND/TRANSPORT chuẩn catalog nếu tháng có phát sinh
        pairs: list[tuple[str, Decimal]] = []
        if attend > 0:
            pairs.append(("ATTEND", Decimal("230000")))
        if trans > 0:
            pairs.append(("TRANSPORT", Decimal("760000")))
        if position > 0:
            pairs.append(("POSITION", _monthly_from_prorata(position, total_d)))
        if toxic > 0:
            pairs.append(("TOXIC", _monthly_from_prorata(toxic, total_d)))
        if pccc > 0:
            pairs.append(("PCCC", _monthly_from_prorata(pccc, total_d)))
        if tech > 0:
            pairs.append(("TECH", _monthly_from_prorata(tech, total_d)))
        if senior > 0:
            pairs.append(("SENIORITY", senior))  # Excel thường ghi mức tháng
        if other > 0:
            pairs.append(("OTHER", _monthly_from_prorata(other, total_d)))

        for ac, amt in pairs:
            r = _upsert_allowance(db, emp, types, ac, amt)
            if r == "created":
                asg_created += 1
            elif r == "updated":
                asg_updated += 1

        ts = (
            db.query(TimesheetMonth)
            .filter(
                TimesheetMonth.pay_period_id == pay.id,
                TimesheetMonth.employee_id == emp.id,
            )
            .one_or_none()
        )
        if ts is None:
            ts = TimesheetMonth(pay_period_id=pay.id, employee_id=emp.id)
            db.add(ts)
        ts.worked_days = total_d
        ts.al_days = al
        ts.rem_days = rem
        ts.ot_hours_weekday = ot_h
        ts.ot_hours_weekend = ZERO
        ts.ot_hours_holiday = ZERO
        ts.late_count = 0
        ts.early_count = 0
        ts_upserted += 1

        if (emp_created + emp_updated) % 50 == 0:
            db.flush()

    db.commit()
    return {
        "period": PERIOD,
        "employees_created": emp_created,
        "employees_updated": emp_updated,
        "allowances_created": asg_created,
        "allowances_updated": asg_updated,
        "timesheets": ts_upserted,
        "skipped_rows": skipped,
        "source": str(path),
    }


def main() -> None:
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/salary_dec.xlsx")
    if not src.is_file():
        raise SystemExit(f"Không thấy file: {src}")
    db: Session = SessionLocal()
    try:
        result = load(db, src)
        print("Trợ Lý AI: nạp Salary Dec.2025 xong:", result)
        print("Gợi ý: Portal → Tính Lương → kỳ 2025-12 → Tính lương")
    finally:
        db.close()


if __name__ == "__main__":
    main()
