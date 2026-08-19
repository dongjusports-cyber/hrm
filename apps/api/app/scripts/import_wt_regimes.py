"""Nạp danh sách chế độ thai sản / nuôi con nhỏ từ Excel 18.08.

Cho phép date_from trong quá khứ (đường import, không qua API HR).

  python -m app.scripts.import_wt_regimes
  python -m app.scripts.import_wt_regimes --xlsx "Dữ liệu nhân viên/Dữ liệu công nhân/Chế độ thai sản nuôi con nhỏ18.08.xlsx"
  python -m app.scripts.import_wt_regimes --dry-run
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.modules.attendance.models import WorkShift  # noqa: F401
from app.modules.core.models import User
from app.modules.mdm.models import Employee, EmployeeWtRegime
from app.modules.mdm.service import _cut_open_regimes_before, _recalc_after_regime_change


def _default_xlsx() -> Path:
    here = Path(__file__).resolve()
    name = "Chế độ thai sản nuôi con nhỏ18.08.xlsx"
    for parent in here.parents:
        cand = parent / "Dữ liệu nhân viên" / "Dữ liệu công nhân" / name
        if cand.is_file():
            return cand
    return Path("/tmp/che_do.xlsx")

SECTION = {
    "con nhỏ": "CHILD",
    "nuôi con": "CHILD",
    "mang thai": "PREGNANT",
    "thai sản": "PREGNANT",
    "nghỉ thai sản": "MATERNITY",
    "nghỉ đẻ": "MATERNITY",
    "đang nghỉ chế độ đặc biệt ( nghỉ sau khi sanh em bé)": "MATERNITY",
}


def _heading_key(text: str) -> str:
    return " ".join(text.casefold().split())


def map_section_heading(text: str) -> str | None:
    """Nhận tiêu đề mục Excel → PREGNANT / MATERNITY / CHILD."""
    key = _heading_key(text)
    if key in SECTION:
        return SECTION[key]
    if "nghỉ sau khi sanh" in key or key.startswith("đang nghỉ chế độ"):
        return "MATERNITY"
    return None
DEFAULT_HOURS = {"PREGNANT": 1, "CHILD": 2, "MATERNITY": 0}
LINE_RE = re.compile(
    r"^(\d+)\s*-\s*(.+?)\s*->\s*(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})\s*$"
)


@dataclass(frozen=True)
class WtRegimeRow:
    employee_code: str
    full_name: str
    regime_type: str
    hours_early: int
    date_from: date
    date_to: date
    note: str


def _parse_date(raw: str) -> date:
    return datetime.strptime(raw.strip(), "%d/%m/%Y").date()


def parse_wt_regime_lines(texts: list[str]) -> list[WtRegimeRow]:
    """Parse dòng Excel: '6472 - TÊN - bộ phận -> dd/mm/yyyy - dd/mm/yyyy'."""
    regime_type: str | None = None
    out: list[WtRegimeRow] = []
    for raw in texts:
        text = (raw or "").strip()
        if not text:
            continue
        key = _heading_key(text)
        if key.startswith("danh sách"):
            continue
        mapped = map_section_heading(text)
        if mapped:
            regime_type = mapped
            continue
        m = LINE_RE.match(text)
        if m is None:
            raise ValueError(f"Không đọc được dòng chế độ: {text}")
        if regime_type is None:
            raise ValueError(f"Dòng NV trước khi có mục Con nhỏ / Mang thai / Nghỉ sau sanh: {text}")
        code, blob, d1, d2 = m.groups()
        name = blob.split(" - ")[0].strip()
        date_from = _parse_date(d1)
        date_to = _parse_date(d2)
        if date_to < date_from:
            raise ValueError(f"MSNV {code}: ngày kết thúc trước ngày bắt đầu.")
        out.append(
            WtRegimeRow(
                employee_code=code,
                full_name=name,
                regime_type=regime_type,
                hours_early=DEFAULT_HOURS[regime_type],
                date_from=date_from,
                date_to=date_to,
                note=f"Nạp Excel 18.08 — {text}",
            )
        )
    return out


def load_wt_regime_xlsx(path: Path) -> list[WtRegimeRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        texts: list[str] = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = row[0]
            if val is not None and str(val).strip():
                texts.append(str(val).strip())
    finally:
        wb.close()
    return parse_wt_regime_lines(texts)


def _same_regime(existing: EmployeeWtRegime, row: WtRegimeRow) -> bool:
    return (
        existing.regime_type == row.regime_type
        and existing.date_from == row.date_from
        and existing.date_to == row.date_to
        and existing.hours_early == row.hours_early
        and existing.ended_at is None
    )


def run(xlsx: Path, *, dry_run: bool = False) -> None:
    rows = load_wt_regime_xlsx(xlsx)
    print(f"Đọc {len(rows)} dòng từ {xlsx.name}")
    db = SessionLocal()
    added = skipped = missing = cut = 0
    try:
        actor = db.query(User).filter(User.username == "admin").first()
        actor_id = actor.id if actor is not None else None
        for row in rows:
            emp = (
                db.query(Employee)
                .filter(Employee.employee_code == row.employee_code, Employee.deleted_at.is_(None))
                .first()
            )
            if emp is None:
                print(f"  thiếu NV {row.employee_code} {row.full_name}")
                missing += 1
                continue
            if row.regime_type == "MATERNITY" and emp.status == "resigned":
                emp.status = "maternity"
                emp.resign_date = None
            existing = (
                db.query(EmployeeWtRegime)
                .filter(EmployeeWtRegime.employee_id == emp.id, EmployeeWtRegime.ended_at.is_(None))
                .order_by(EmployeeWtRegime.date_from.desc())
                .all()
            )
            if any(_same_regime(r, row) for r in existing):
                print(
                    f"  bỏ qua {row.employee_code} {row.regime_type} "
                    f"{row.date_from.isoformat()}–{row.date_to.isoformat()} (đã có)"
                )
                skipped += 1
                continue
            cut_rows = _cut_open_regimes_before(db, emp.id, row.date_from)
            if cut_rows:
                cut += 1
            rec = EmployeeWtRegime(
                employee_id=emp.id,
                regime_type=row.regime_type,
                hours_early=row.hours_early,
                date_from=row.date_from,
                date_to=row.date_to,
                note=row.note[:500],
                created_by_user_id=actor_id,
            )
            db.add(rec)
            db.flush()
            if not dry_run:
                for old in cut_rows:
                    _recalc_after_regime_change(db, emp, old)
                _recalc_after_regime_change(db, emp, rec)
            kind = {
                "CHILD": "Nuôi con nhỏ",
                "MATERNITY": "Nghỉ thai sản",
                "PREGNANT": "Đang mang thai",
            }.get(row.regime_type, row.regime_type)
            hours_bit = "" if row.regime_type == "MATERNITY" else f" {row.hours_early}h"
            print(
                f"  + {row.employee_code} {row.full_name} — {kind}{hours_bit} "
                f"{row.date_from.strftime('%d/%m/%Y')}–{row.date_to.strftime('%d/%m/%Y')}"
            )
            added += 1
        if dry_run:
            db.rollback()
            print("(dry-run) không ghi DB.")
        else:
            db.commit()
        print(
            f"Thêm {added} | đã có {skipped} | cắt giai đoạn cũ {cut} | không có trong DB {missing}"
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", type=Path, default=None)
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()
    xlsx = args.xlsx or _default_xlsx()
    if not xlsx.is_file():
        raise SystemExit(f"Không thấy file Excel: {xlsx}")
    run(xlsx, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
