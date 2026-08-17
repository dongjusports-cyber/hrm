"""Nạp hồ sơ + đồng bộ Bộ phận › Tổ từ 3 file HIEN_PHAP (11.08.2026).

Nguồn:
  - Bộ phận_11.08.xlsx          → danh sách 30 tổ còn hiệu lực (tháng 7)
  - Danh sách Nhân Viên - Bộ phận hiện tại.xls → MSNV + DEPARTMENT + GROUP (ánh xạ tổ)
  - Thông tin danh sách công nhân 11.08.26.xls   → hồ sơ chi tiết (ngày sinh, ĐC, SĐT…)

Chạy:
  docker cp "C:/DATA/HRM/dj-hrm/dj-hrm/HIEN_PHAP/Thông tin danh sách nhân viên" djhrm-api:/tmp/empinfo
  docker compose exec api python -m app.scripts.import_employee_list_1108 /tmp/empinfo --profiles-only [--dry-run]

  # Full (đồng bộ tổ — cẩn thận trên DB công ty):
  docker compose exec api python -m app.scripts.import_employee_list_1108 /tmp/empinfo
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.modules.attendance.models import WorkShift  # noqa: F401 — FK teams.default_shift_id
from app.modules.audit.models import AuditLog
from app.modules.core.models import User  # noqa: F401 — FK audit_logs.actor_user_id
from app.modules.mdm.models import Department, Employee, LabourContract, Team
from app.scripts.empinfo_lookup_map import (
    infer_contract_type,
    resolve_education_code,
    resolve_nationality_code,
    resolve_place_code,
)

ORG_EFFECTIVE = date(2026, 7, 1)
CLOSE_DATE = date(2026, 6, 30)


def _as_decimal(val, default: Decimal = Decimal("0")) -> Decimal:
    if val is None or val == "":
        return default
    if isinstance(val, Decimal):
        return val
    return Decimal(str(val))

TEAM_ALIASES: dict[str, str] = {
    "HR/Admin": "HR/Admin Staff",
    "QC (KCS)": "QC (KCS)",
}

DEPT_ALIASES: dict[str, str] = {
    "quality management": "Quality management",
    "hr & admin": "HR & Admin",
    "main office": "Main Office",
}


def _norm(s: str) -> str:
    return unicodedata.normalize("NFC", (s or "").strip())


def _parse_vn_date(raw, book_datemode: int = 0) -> date | None:
    if raw in ("", None):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        try:
            return xlrd.xldate_as_datetime(raw, book_datemode).date()
        except Exception:
            return None
    text = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _clean_digits(raw) -> str | None:
    if raw in ("", None):
        return None
    s = re.sub(r"\D", "", str(raw).strip())
    return s or None


def _clean_phone(raw) -> str | None:
    if raw in ("", None):
        return None
    if isinstance(raw, float):
        if raw != int(raw):
            return None
        raw = int(raw)
    s = re.sub(r"\s", "", str(raw))
    s = re.sub(r"\D", "", s)
    # Excel hay lưu 908275468 (mất số 0 đầu) — SĐT VN 10 số.
    if s and not s.startswith("0") and len(s) in (9, 10):
        s = "0" + s
    if len(s) < 9 or len(s) > 11:
        return None
    if not s.startswith("0"):
        return None
    return s


def _gender(raw: str | None) -> str | None:
    if not raw:
        return None
    t = _norm(raw).lower()
    if t in ("male", "m", "nam"):
        return "M"
    if t in ("female", "f") or t.startswith("n"):
        return "F"
    return None


def _education_code(raw) -> str | None:
    return resolve_education_code(raw)


def load_valid_teams(xlsx_path: Path) -> set[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    names: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            name = _norm(str(row[0]))
            if name and name.lower() != "position bộ phận":
                names.add(name)
    wb.close()
    return names


def load_assignments(xls_path: Path) -> dict[str, dict]:
    """File 2 — DEPARTMENT + GROUP theo MSNV."""
    wb = xlrd.open_workbook(str(xls_path))
    sh = wb.sheet_by_index(0)
    out: dict[str, dict] = {}
    for r in range(3, sh.nrows):
        try:
            msnv = str(int(float(sh.cell_value(r, 3))))
        except (TypeError, ValueError):
            continue
        dept = _norm(str(sh.cell_value(r, 1)))
        team = _norm(str(sh.cell_value(r, 2)))
        if not dept or not team:
            continue
        out[msnv] = {
            "department_name": dept,
            "team_name": team,
            "full_name": _norm(str(sh.cell_value(r, 4))),
            "join_date": _parse_vn_date(sh.cell_value(r, 5), wb.datemode),
            "position_title": _norm(str(sh.cell_value(r, 6))) or None,
            "contract_salary": Decimal(str(sh.cell_value(r, 8) or 0)),
        }
    return out


def load_profiles(xls_path: Path) -> dict[str, dict]:
    """File 3 — hồ sơ chi tiết."""
    wb = xlrd.open_workbook(str(xls_path))
    sh = wb.sheet_by_index(0)
    dm = wb.datemode
    out: dict[str, dict] = {}
    for r in range(3, sh.nrows):
        try:
            msnv = str(int(float(sh.cell_value(r, 1))))
        except (TypeError, ValueError):
            continue
        dept = _norm(str(sh.cell_value(r, 0)))
        team = _norm(str(sh.cell_value(r, 8)))
        if dept.lower() in ("dept", ""):
            continue
        left = _parse_vn_date(sh.cell_value(r, 21), dm)
        contract_no = _norm(str(sh.cell_value(r, 18))) or None
        out[msnv] = {
            "department_name": dept,
            "team_name": team,
            "full_name": _norm(str(sh.cell_value(r, 2))),
            "birth_date": _parse_vn_date(sh.cell_value(r, 4), dm),
            "join_date": _parse_vn_date(sh.cell_value(r, 5), dm),
            "birth_place_text": _norm(str(sh.cell_value(r, 6))) or None,
            "contract_salary": Decimal(str(sh.cell_value(r, 7) or 0)),
            "gender": _gender(str(sh.cell_value(r, 9))),
            "position_title": _norm(str(sh.cell_value(r, 10))) or None,
            "phone": _clean_phone(sh.cell_value(r, 11)),
            "education_code": _education_code(sh.cell_value(r, 12)),
            "permanent_address": _norm(str(sh.cell_value(r, 13)).replace("\xa0", " ")) or None,
            "id_number": _clean_digits(sh.cell_value(r, 14)),
            "id_issue_place_text": _norm(str(sh.cell_value(r, 15))) or None,
            "id_issue_date": _parse_vn_date(sh.cell_value(r, 16), dm),
            "si_book_no": _clean_digits(sh.cell_value(r, 17)),
            "contract_no": contract_no,
            "contract_type_code": infer_contract_type(contract_no),
            "contract_start": _parse_vn_date(sh.cell_value(r, 19), dm),
            "contract_end": _parse_vn_date(sh.cell_value(r, 20), dm),
            "left_date": left,
            "status": "resigned" if left else "active",
        }
    return out


def _dept_by_name(db: Session, name: str) -> Department | None:
    key = _norm(name)
    alias = DEPT_ALIASES.get(key.lower(), key)
    row = db.query(Department).filter(Department.name == alias).first()
    if row:
        return row
    return db.query(Department).filter(Department.name.ilike(alias)).first()


def _team_names_match(a: str, b: str) -> bool:
    a, b = _norm(a), _norm(b)
    if a == b:
        return True
    if TEAM_ALIASES.get(a) == b or TEAM_ALIASES.get(b) == a:
        return True
    return False


def _find_team(db: Session, dept: Department, team_name: str) -> Team | None:
    name = _norm(team_name)
    teams = db.query(Team).filter(Team.department_id == dept.id).all()
    for t in teams:
        if _team_names_match(t.name, name):
            return t
    return None


def _next_team_code(db: Session, dept_id) -> str:
    codes = [
        int(t.code)
        for t in db.query(Team).filter(Team.department_id == dept_id).all()
        if str(t.code).isdigit()
    ]
    return str(max(codes or [0]) + 1).zfill(2)


def ensure_team(db: Session, dept: Department, team_name: str, *, valid_teams: set[str]) -> Team | None:
    name = _norm(team_name)
    canonical = TEAM_ALIASES.get(name, name)
    if canonical not in valid_teams and name not in valid_teams:
        return None
    use_name = canonical if canonical in valid_teams else name
    existing = _find_team(db, dept, use_name)
    if existing:
        existing.name = use_name
        existing.name_local = use_name
        existing.effective_to = None
        if existing.effective_from > ORG_EFFECTIVE:
            existing.effective_from = ORG_EFFECTIVE
        return existing
    team = Team(
        department_id=dept.id,
        code=_next_team_code(db, dept.id),
        name=use_name,
        name_local=use_name,
        effective_from=ORG_EFFECTIVE,
        effective_to=None,
    )
    db.add(team)
    db.flush()
    return team


def sync_org(
    db: Session,
    pairs: set[tuple[str, str]],
    valid_teams: set[str],
    *,
    dry_run: bool,
) -> dict[tuple[str, str], Team]:
    """Đóng tổ không còn trong danh sách; mở/ tạo tổ mới đúng Bộ phận."""
    team_map: dict[tuple[str, str], Team] = {}
    for dept_name, team_name in sorted(pairs):
        dept = _dept_by_name(db, dept_name)
        if dept is None:
            print(f"  ! Không tìm thấy bộ phận: {dept_name}")
            continue
        if dry_run:
            team_map[(dept_name, team_name)] = None  # type: ignore[assignment]
            continue
        team = ensure_team(db, dept, team_name, valid_teams=valid_teams)
        if team:
            team_map[(dept_name, team_name)] = team

    if not dry_run:
        closed = 0
        deleted = 0
        for team in db.query(Team).all():
            if team.name in valid_teams or any(
                _team_names_match(team.name, v) for v in valid_teams
            ):
                continue
            n_emp = db.query(Employee).filter(Employee.team_id == team.id).count()
            if n_emp == 0 and team.effective_to is not None:
                db.delete(team)
                deleted += 1
                continue
            if team.effective_to is None:
                team.effective_to = CLOSE_DATE
                closed += 1
        for dept in db.query(Department).all():
            active_teams = [
                t
                for t in dept.teams
                if t.effective_to is None
                and (
                    t.name in valid_teams
                    or any(_team_names_match(t.name, v) for v in valid_teams)
                )
            ]
            if not active_teams and dept.name not in {
                "Newcomers",
                "PURCHASE",
            }:
                if dept.effective_to is None:
                    dept.effective_to = CLOSE_DATE
        db.flush()
        print(f"  Đóng {closed} tổ cũ; xóa {deleted} tổ không còn NV.")

    return team_map


def _apply_profile(emp: Employee, assign: dict, prof: dict) -> None:
    """Ghi đè/bổ sung hồ sơ từ Excel — tên luôn lấy file 3 nếu có (UTF-8 đúng)."""
    if prof.get("full_name"):
        emp.full_name = prof["full_name"]
    elif assign.get("full_name"):
        emp.full_name = assign["full_name"]

    for field in (
        "birth_date",
        "join_date",
        "gender",
        "phone",
        "education_code",
        "permanent_address",
        "temporary_address",
        "id_number",
        "id_issue_date",
        "si_book_no",
        "bank_account",
        "marital_status",
        "pay_channel",
    ):
        val = prof.get(field)
        if val is not None and val != "":
            setattr(emp, field, val)
    if prof.get("children_count") is not None:
        emp.children_count = int(prof["children_count"] or 0)

    birth_code = resolve_place_code("birth_place", prof.get("birth_place_text"))
    if birth_code:
        emp.birth_place_code = birth_code
    issue_code = resolve_place_code("id_issue_place", prof.get("id_issue_place_text"))
    if issue_code:
        emp.id_issue_place_code = issue_code

    # Quốc tịch: mặc định VN khi có CCCD/CMND Việt hoặc có hồ sơ trong file
    if emp.id_number or emp.birth_place_code or prof:
        emp.nationality_code = resolve_nationality_code(None)

    if assign.get("join_date") and not emp.join_date:
        emp.join_date = assign["join_date"]
    if assign.get("position_title"):
        emp.position_title = assign["position_title"]
    elif prof.get("position_title"):
        emp.position_title = prof["position_title"]

    sal = prof.get("contract_salary") or assign.get("contract_salary")
    sal_dec = _as_decimal(sal, Decimal("-1"))
    if sal_dec > 0:
        emp.contract_salary = sal_dec

    signed = prof.get("contract_signed_at") or prof.get("contract_start")
    if signed:
        emp.contract_signed_at = signed if isinstance(signed, date) else _parse_vn_date(signed, 0)

    ctype = prof.get("contract_type_code")
    if prof.get("left_date"):
        emp.resign_date = prof["left_date"]
        emp.status = "resigned"
    elif ctype == "TV":
        emp.status = "probation"
    elif ctype in ("VTH", "HD1", "HD2"):
        emp.status = "active"
        emp.resign_date = None
    else:
        emp.status = "active"
        emp.resign_date = None


def _upsert_labour_contract(db: Session, emp: Employee, prof: dict) -> None:
    """Cập nhật HĐ đang active từ cột Contract No / Start / End — không tạo lịch sử giả."""
    start = prof.get("contract_start") or emp.contract_signed_at or emp.join_date
    if not start:
        return
    ctype = prof.get("contract_type_code") or "VTH"
    end = prof.get("contract_end")
    salary = _as_decimal(prof.get("contract_salary") or emp.contract_salary)

    active = (
        db.query(LabourContract)
        .filter(LabourContract.employee_id == emp.id, LabourContract.status == "active")
        .order_by(LabourContract.start_date.desc())
        .first()
    )
    if active is None:
        active = LabourContract(
            employee_id=emp.id,
            contract_type_code=ctype,
            seq_no=1,
            sign_date=start,
            start_date=start,
            end_date=end,
            base_salary=salary,
            team_id=emp.team_id,
            status="active" if emp.status != "resigned" else "terminated",
        )
        db.add(active)
        return

    active.contract_type_code = ctype
    active.sign_date = start
    active.start_date = start
    if end is not None:
        active.end_date = end
    elif ctype == "VTH":
        active.end_date = None
    if salary > 0:
        active.base_salary = salary
    if emp.team_id:
        active.team_id = emp.team_id
    if emp.status == "resigned":
        active.status = "terminated"
        if emp.resign_date and active.end_date is None:
            active.end_date = emp.resign_date


def run(
    data_dir: str,
    *,
    dry_run: bool = False,
    profiles_only: bool = False,
) -> None:
    base = Path(data_dir)
    xlsx = next(base.glob("Bộ phận*.xlsx"))
    xls_assign = next(base.glob("Danh sách*.xls"))
    xls_profile = next(base.glob("Thông tin*.xls"))

    valid_teams = load_valid_teams(xlsx)
    assignments = load_assignments(xls_assign)
    profiles = load_profiles(xls_profile)

    print(f"Tổ hợp lệ (xlsx): {len(valid_teams)}")
    print(f"NV file gán tổ: {len(assignments)} | file hồ sơ: {len(profiles)}")
    print(f"Chế độ: {'profiles-only (không đụng cây tổ)' if profiles_only else 'full (đồng bộ tổ)'}")

    active_codes = set(assignments.keys()) | set(profiles.keys())
    org_pairs: set[tuple[str, str]] = set()
    for rec in assignments.values():
        org_pairs.add((rec["department_name"], rec["team_name"]))
    for rec in profiles.values():
        if rec.get("department_name") and rec.get("team_name"):
            org_pairs.add((rec["department_name"], rec["team_name"]))

    db = SessionLocal()
    try:
        if not profiles_only:
            print("== Đồng bộ cây tổ chức ==")
            sync_org(db, org_pairs, valid_teams, dry_run=dry_run)

        updated = created = resigned = no_team = 0
        place_ok = edu_ok = 0
        for msnv in sorted(active_codes):
            assign = assignments.get(msnv, {})
            prof = profiles.get(msnv, {})
            dept_name = assign.get("department_name") or prof.get("department_name")
            team_name = assign.get("team_name") or prof.get("team_name")

            emp = db.query(Employee).filter(Employee.employee_code == msnv).first()
            is_new = emp is None
            if is_new:
                emp = Employee(employee_code=msnv, status="active")
                db.add(emp)
                created += 1
            else:
                updated += 1
                emp.deleted_at = None

            _apply_profile(emp, assign, prof)
            if emp.birth_place_code:
                place_ok += 1
            if emp.education_code and str(emp.education_code).startswith("EDUCATION_"):
                edu_ok += 1

            if dry_run:
                continue

            if not profiles_only:
                dept = _dept_by_name(db, dept_name) if dept_name else None
                team = None
                if dept and team_name:
                    team = ensure_team(db, dept, team_name, valid_teams=valid_teams)
                if team:
                    emp.team_id = team.id
                else:
                    no_team += 1

            db.flush()
            if prof:
                _upsert_labour_contract(db, emp, prof)

        if not dry_run:
            if not profiles_only:
                for emp in db.query(Employee).filter(Employee.deleted_at.is_(None)).all():
                    if emp.employee_code not in active_codes and emp.status != "resigned":
                        emp.status = "resigned"
                        if not emp.resign_date:
                            emp.resign_date = ORG_EFFECTIVE
                        resigned += 1

            db.add(
                AuditLog(
                    actor_username="system.import_employee_list_1108",
                    action="import_employee_list",
                    entity_type="employee",
                    entity_id="bulk",
                    summary=(
                        f"Nạp hồ sơ 11.08.2026 ({'profiles-only' if profiles_only else 'full'}): "
                        f"cập nhật {updated}, mới {created}, thôi việc {resigned}, "
                        f"nơi sinh {place_ok}, học vấn {edu_ok}."
                    ),
                )
            )
            db.commit()

        print(
            f"HOÀN TẤT: cập nhật {updated}, mới {created}, "
            f"đánh dấu nghỉ {resigned}, chưa gán tổ {no_team}."
        )
        print(f"  Nơi sinh map được: {place_ok} | Học vấn EDUCATION_*: {edu_ok}")
        if dry_run:
            print("(dry-run) không ghi DB.")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("data_dir", help="Thư mục chứa 3 file Excel")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--profiles-only",
        action="store_true",
        help="Chỉ bổ sung hồ sơ/HĐ — không đóng/mở tổ, không đổi team_id",
    )
    args = parser.parse_args()
    run(args.data_dir, dry_run=args.dry_run, profiles_only=args.profiles_only)


if __name__ == "__main__":
    main()
