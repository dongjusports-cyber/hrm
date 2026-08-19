"""Parse + nạp file «Thông tin bổ sung 14.08.26.xlsx» vào hồ sơ công nhân.

Sheet:
  «giới tính- hôn nhân» — giới tính (0 nữ / 1 nam), hôn nhân (0 độc thân, -1/1 đã kết hôn), số con
  «02» — hồ sơ; cột M (Education) = trình độ
  «03» — phụ cấp cố định (PCCC+HSE, chức vụ, độc hại, chuyên cần, đi lại, tay nghề, khác)
  «01» — bảng lương T8/2026 (không nạp vào hồ sơ)

Mặc định toàn bộ NV: quốc tịch Việt Nam, dân tộc Kinh, tôn giáo Không.

Chạy:
  python -m app.scripts.extract_thong_tin_bo_sung
  python -m app.scripts.import_thong_tin_bo_sung [--dry-run]
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.scripts.empinfo_lookup_map import (
    infer_contract_type,
    resolve_education_code,
    resolve_nationality_code,
    resolve_place_code,
)
from app.scripts.import_employee_list_1108 import (
    _clean_digits,
    _clean_phone,
    _gender,
    _norm,
    _parse_vn_date,
)
from app.modules.payroll.seed_allowances import normalize_legacy_allowance_amount

SKIP_TEST_CODES = {"1604", "1718", "8851"}

DEFAULT_NATIONALITY = "NATIONALITY001"  # Việt Nam
DEFAULT_ETHNICITY = "ETHNICITY001"  # Kinh
DEFAULT_RELIGION = "RELIGION001"  # Không

HSE_DEFAULT = Decimal("50000")
PCCC_HSE_COMBINED = Decimal("932000")
MANAGED_ALLOWANCE_CODES = (
    "ATTEND",
    "TRANSPORT",
    "POSITION",
    "TOXIC",
    "PCCC",
    "HSE",
    "TECH",
    "OTHER",
)
SHEET03_FIXED_COLS: list[tuple[int, str]] = [
    (10, "POSITION"),
    (11, "TOXIC"),
    (12, "ATTEND"),
    (13, "TRANSPORT"),
    (14, "TECH"),
    (15, "OTHER"),
]

# VBA/Excel Boolean True = -1; header ghi 1 = đã kết hôn.
_MARRIED_FLAGS = {-1, 1, True, "1", "-1"}
_SINGLE_FLAGS = {0, False, "0"}


def default_xlsx_path() -> Path:
    root = Path(__file__).resolve().parents[4]
    folder = root / "Dữ liệu nhân viên" / "Dữ liệu công nhân"
    matches = sorted(
        p
        for p in list(folder.glob("*bổ sung*.xlsx")) + list(folder.glob("*bo sung*.xlsx"))
        if not p.name.startswith("~$")
    )
    if not matches:
        matches = [p for p in folder.glob("*.xlsx") if "b" in p.name.lower()]
    if not matches:
        raise FileNotFoundError(f"Không thấy file Thông tin bổ sung .xlsx trong {folder}")
    return matches[0]


def default_snapshot_dir() -> Path:
    return Path(__file__).resolve().parents[4] / "Dữ liệu nhân viên" / "Dữ liệu công nhân"


def default_extract_dir(xlsx_path: Path | None = None) -> Path:
    src = xlsx_path or default_xlsx_path()
    return src.parent / "trich_xuat_bo_sung_140826"


def _msnv(raw: Any) -> str | None:
    if raw in ("", None):
        return None
    if isinstance(raw, float):
        if raw != int(raw):
            return None
        return str(int(raw))
    if isinstance(raw, int):
        return str(raw)
    text = str(raw).strip()
    if not text:
        return None
    try:
        return str(int(float(text)))
    except ValueError:
        digits = re.sub(r"\D", "", text)
        return digits or None


def _as_date(raw: Any) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return _parse_vn_date(raw, 0)


def _as_decimal(raw: Any) -> Decimal | None:
    if raw in ("", None):
        return None
    try:
        n = Decimal(str(raw).strip().replace(",", ""))
    except Exception:
        return None
    return n if n > 0 else None


def _allowance_amount(raw: Any) -> Decimal:
    if raw in ("", None):
        return Decimal("0")
    try:
        n = Decimal(str(raw).strip().replace(",", ""))
    except Exception:
        return Decimal("0")
    return n if n > 0 else Decimal("0")


def split_pccc_hse(amt: Decimal) -> dict[str, Decimal]:
    """Sheet 03 cột PCCC+HSE_AMT → mã PCCC / HSE riêng (quy ước CTY 2026-08)."""
    if amt <= 0:
        return {}
    if amt == HSE_DEFAULT:
        return {"HSE": HSE_DEFAULT}
    if amt == PCCC_HSE_COMBINED:
        return {"HSE": HSE_DEFAULT, "PCCC": amt - HSE_DEFAULT}
    return {"PCCC": amt}


def parse_sheet03_allowances(row: tuple[Any, ...]) -> dict[str, Decimal]:
    amounts = {code: Decimal("0") for code in MANAGED_ALLOWANCE_CODES}
    for col, code in SHEET03_FIXED_COLS:
        raw = row[col] if len(row) > col else None
        amt = _allowance_amount(raw)
        if code in ("ATTEND", "TRANSPORT"):
            amt = normalize_legacy_allowance_amount(code, amt) or amt
        amounts[code] = amt
    split = split_pccc_hse(_allowance_amount(row[9] if len(row) > 9 else None))
    amounts["PCCC"] = split.get("PCCC", Decimal("0"))
    amounts["HSE"] = split.get("HSE", Decimal("0"))
    return amounts


def allowances_as_json(amounts: dict[str, Decimal]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for code in MANAGED_ALLOWANCE_CODES:
        amt = amounts.get(code) or Decimal("0")
        if amt > 0:
            items.append({"component_code": code, "amount": f"{amt:.2f}", "meta": None})
    return items


def apply_default_identity(dest: dict[str, Any]) -> list[str]:
    changed: list[str] = []
    for field, val in (
        ("nationality_code", DEFAULT_NATIONALITY),
        ("ethnicity_code", DEFAULT_ETHNICITY),
        ("religion_code", DEFAULT_RELIGION),
    ):
        if dest.get(field) != val:
            dest[field] = val
            changed.append(field)
    return changed


def resolve_marital_status(raw: Any) -> str | None:
    """0 = độc thân; -1 hoặc 1 = đã kết hôn (Excel/VBA True = -1)."""
    if raw in ("", None):
        return None
    if raw in _MARRIED_FLAGS:
        return "married"
    if raw in _SINGLE_FLAGS:
        return "single"
    text = _norm(str(raw)).lower()
    if text in ("married", "đã kết hôn", "da ket hon", "kết hôn", "ket hon"):
        return "married"
    if text in ("single", "độc thân", "doc than"):
        return "single"
    try:
        n = int(float(str(raw).strip()))
    except ValueError:
        return None
    if n in _MARRIED_FLAGS:
        return "married"
    if n in _SINGLE_FLAGS:
        return "single"
    return None


def resolve_gender_flag(raw: Any) -> str | None:
    """Sheet hôn nhân: 0 = nữ / 1 = nam. Trả M/F (khớp import GenusSuite)."""
    if raw in ("", None):
        return None
    if raw in (0, "0", False):
        return "F"
    if raw in (1, "1", True, -1):
        return "M"
    return _gender(str(raw))


def _children_count(raw: Any) -> int:
    if raw in ("", None):
        return 0
    try:
        n = int(float(str(raw).strip()))
    except ValueError:
        return 0
    return max(0, n)


def _bank_account(raw: Any, phone: str | None) -> str | None:
    digits = _clean_digits(raw)
    if not digits or len(digits) < 8:
        return None
    if phone and digits == phone:
        return None
    if phone and digits.lstrip("0") == phone.lstrip("0"):
        return None
    return digits


def _prefer_longer_id(current: str | None, incoming: str | None) -> str | None:
    """Excel số BHXH hay mất chữ số cuối — giữ bản dài hơn nếu là tiền tố."""
    cur = (current or "").strip() or None
    inc = (incoming or "").strip() or None
    if not inc:
        return cur
    if not cur:
        return inc
    if cur == inc:
        return cur
    if cur.startswith(inc) and len(cur) > len(inc):
        return cur
    if inc.startswith(cur) and len(inc) > len(cur):
        return inc
    return cur


def _json_default(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    raise TypeError(f"Không serialize được: {type(value)!r}")


def load_supplement(xlsx_path: Path) -> dict[str, dict[str, Any]]:
    """Gộp sheet hôn nhân + sheet 02 theo MSNV."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        by_code: dict[str, dict[str, Any]] = {}

        marital_sheet = None
        for name in wb.sheetnames:
            folded = name.casefold()
            if "hôn nhân" in name or "hon nhan" in folded or "giới tính" in name:
                marital_sheet = wb[name]
                break
        if marital_sheet is not None:
            for i, row in enumerate(marital_sheet.iter_rows(values_only=True)):
                if i < 2:
                    continue
                code = _msnv(row[0] if row else None)
                if not code or code in SKIP_TEST_CODES:
                    continue
                rec = by_code.setdefault(code, {"employee_code": code})
                rec["full_name"] = _norm(str(row[1] or "")) or rec.get("full_name")
                rec["gender"] = resolve_gender_flag(row[2] if len(row) > 2 else None)
                rec["marital_status"] = resolve_marital_status(row[3] if len(row) > 3 else None)
                rec["children_count"] = _children_count(row[4] if len(row) > 4 else None)

        if "02" in wb.sheetnames:
            for i, row in enumerate(wb["02"].iter_rows(values_only=True)):
                if i < 3:
                    continue
                code = _msnv(row[1] if row else None)
                if not code or code in SKIP_TEST_CODES:
                    continue
                rec = by_code.setdefault(code, {"employee_code": code})
                phone = _clean_phone(row[11] if len(row) > 11 else None)
                bank = _bank_account(row[3] if len(row) > 3 else None, phone)
                rec.update(
                    {
                        "department_name": _norm(str(row[0] or "")) or rec.get("department_name"),
                        "full_name": _norm(str(row[2] or "")) or rec.get("full_name"),
                        "bank_account": bank,
                        "birth_date": _as_date(row[4] if len(row) > 4 else None),
                        "join_date": _as_date(row[5] if len(row) > 5 else None),
                        "birth_place_text": _norm(str(row[6] or "")) or None,
                        "contract_salary": _as_decimal(row[7] if len(row) > 7 else None),
                        "team_name": _norm(str(row[8] or "")) or rec.get("team_name"),
                        "position_title": _norm(str(row[10] or "")) or rec.get("position_title"),
                        "phone": phone,
                        "education_code": resolve_education_code(row[12] if len(row) > 12 else None),
                        "permanent_address": _norm(str(row[13] or "").replace("\xa0", " ")) or None,
                        "id_number": _clean_digits(row[14] if len(row) > 14 else None),
                        "id_issue_place_text": _norm(str(row[15] or "")) or None,
                        "id_issue_date": _as_date(row[16] if len(row) > 16 else None),
                        "si_book_no": _clean_digits(row[17] if len(row) > 17 else None),
                        "contract_no": _norm(str(row[18] or "")) or None,
                        "contract_start": _as_date(row[19] if len(row) > 19 else None),
                        "contract_end": _as_date(row[20] if len(row) > 20 else None),
                        "left_date": _as_date(row[21] if len(row) > 21 else None),
                    }
                )
                if rec.get("gender") is None:
                    rec["gender"] = _gender(str(row[9] or ""))
                rec["birth_place_code"] = resolve_place_code("birth_place", rec.get("birth_place_text"))
                rec["id_issue_place_code"] = resolve_place_code(
                    "id_issue_place", rec.get("id_issue_place_text")
                )
                rec["contract_type_code"] = infer_contract_type(rec.get("contract_no"))
                if rec.get("id_number"):
                    rec["nationality_code"] = resolve_nationality_code(None)
                if rec.get("bank_account"):
                    rec["pay_channel"] = "ATM"
                if rec.get("permanent_address"):
                    rec["temporary_address"] = rec["permanent_address"]
                if rec.get("contract_start"):
                    rec["contract_signed_at"] = rec["contract_start"]
                rec["nationality_code"] = DEFAULT_NATIONALITY
                rec["ethnicity_code"] = DEFAULT_ETHNICITY
                rec["religion_code"] = DEFAULT_RELIGION

        if "03" in wb.sheetnames:
            for i, row in enumerate(wb["03"].iter_rows(values_only=True)):
                if i < 3:
                    continue
                code = _msnv(row[3] if row else None)
                if not code or code in SKIP_TEST_CODES:
                    continue
                rec = by_code.setdefault(code, {"employee_code": code})
                amounts = parse_sheet03_allowances(row)
                rec["allowance_amounts"] = amounts
                rec["allowances"] = allowances_as_json(amounts)
        return by_code
    finally:
        wb.close()


# Trường luôn ghi từ file bổ sung (nguồn mới, chưa có trong trich_xuat cũ).
OVERWRITE_FIELDS = ("gender", "marital_status", "children_count")

# Chỉ điền khi hồ sơ đang trống — tránh đè số BHXH/STK đã đúng bằng bản Excel cắt số.
FILL_IF_EMPTY_FIELDS = (
    "phone",
    "bank_account",
    "permanent_address",
    "temporary_address",
    "id_number",
    "id_issue_date",
    "id_issue_place_code",
    "birth_date",
    "birth_place_code",
    "join_date",
    "pay_channel",
    "position_title",
    "contract_signed_at",
)


def _empty(val: Any) -> bool:
    return val in (None, "", [])


def apply_supplement_to_mapping(dest: dict[str, Any], rec: dict[str, Any]) -> list[str]:
    """Ghi rec vào dict hồ sơ (JSON snapshot). Trả danh sách field đã đổi."""
    changed: list[str] = []
    for field in OVERWRITE_FIELDS:
        if field not in rec:
            continue
        val = rec[field]
        if field == "children_count":
            val = int(val or 0)
        if dest.get(field) != val:
            dest[field] = val
            changed.append(field)

    if rec.get("education_code"):
        if dest.get("education_code") != rec["education_code"]:
            dest["education_code"] = rec["education_code"]
            changed.append("education_code")

    changed.extend(apply_default_identity(dest))

    si = _prefer_longer_id(dest.get("si_book_no"), rec.get("si_book_no"))
    if si and dest.get("si_book_no") != si:
        dest["si_book_no"] = si
        changed.append("si_book_no")

    bank = rec.get("bank_account")
    if bank and _empty(dest.get("bank_account")):
        dest["bank_account"] = bank
        changed.append("bank_account")
        if dest.get("pay_channel") in (None, "", "CASH"):
            dest["pay_channel"] = "ATM"
            changed.append("pay_channel")

    for field in FILL_IF_EMPTY_FIELDS:
        if field in ("bank_account", "pay_channel"):
            continue
        val = rec.get(field)
        if _empty(val):
            continue
        if field.endswith("_date") or field in ("join_date", "contract_signed_at"):
            if isinstance(val, date):
                val = val.isoformat()
        if field == "contract_salary" and isinstance(val, Decimal):
            val = f"{val:.2f}"
        if _empty(dest.get(field)):
            dest[field] = val
            changed.append(field)

    if rec.get("left_date") and _empty(dest.get("resign_date")):
        left = rec["left_date"]
        dest["resign_date"] = left.isoformat() if isinstance(left, date) else left
        dest["status"] = "resigned"
        changed.append("resign_date")

    return changed


def apply_supplement_to_employee(emp: Any, rec: dict[str, Any]) -> list[str]:
    """Ghi rec vào model Employee. Trả field đã đổi."""
    changed: list[str] = []
    for field in OVERWRITE_FIELDS:
        if field not in rec:
            continue
        val = rec[field]
        if field == "children_count":
            val = int(val or 0)
        if getattr(emp, field, None) != val:
            setattr(emp, field, val)
            changed.append(field)

    if rec.get("education_code") and emp.education_code != rec["education_code"]:
        emp.education_code = rec["education_code"]
        changed.append("education_code")
    for field, val in (
        ("nationality_code", DEFAULT_NATIONALITY),
        ("ethnicity_code", DEFAULT_ETHNICITY),
        ("religion_code", DEFAULT_RELIGION),
    ):
        if getattr(emp, field, None) != val:
            setattr(emp, field, val)
            changed.append(field)

    si = _prefer_longer_id(getattr(emp, "si_book_no", None), rec.get("si_book_no"))
    if si and emp.si_book_no != si:
        emp.si_book_no = si
        changed.append("si_book_no")

    bank = rec.get("bank_account")
    if bank and _empty(emp.bank_account):
        emp.bank_account = bank
        changed.append("bank_account")
        if emp.pay_channel in (None, "", "CASH"):
            emp.pay_channel = "ATM"
            changed.append("pay_channel")

    for field in FILL_IF_EMPTY_FIELDS:
        if field in ("bank_account", "pay_channel"):
            continue
        val = rec.get(field)
        if _empty(val):
            continue
        if _empty(getattr(emp, field, None)):
            setattr(emp, field, val)
            changed.append(field)

    if rec.get("left_date") and _empty(emp.resign_date):
        emp.resign_date = rec["left_date"]
        emp.status = "resigned"
        changed.append("resign_date")

    return changed


def merge_into_snapshots(records: dict[str, dict[str, Any]], snapshot_dir: Path) -> dict[str, int]:
    emp_dir = snapshot_dir / "employees"
    stats = {
        "files": 0,
        "updated": 0,
        "missing_json": 0,
        "skipped_test": 0,
        "identity_all": 0,
        "allowances": 0,
    }
    touched: set[str] = set()
    for code, rec in records.items():
        if code in SKIP_TEST_CODES:
            stats["skipped_test"] += 1
            continue
        path = emp_dir / f"{code}.json"
        if not path.is_file():
            stats["missing_json"] += 1
            continue
        snap = json.loads(path.read_text(encoding="utf-8"))
        profile = snap.setdefault("profile", {})
        changed = apply_supplement_to_mapping(profile, rec)
        if rec.get("allowances") is not None:
            existing = snap.get("allowances") or []
            kept = [a for a in existing if a.get("component_code") not in MANAGED_ALLOWANCE_CODES]
            incoming = rec["allowances"]
            if kept + incoming != existing:
                snap["allowances"] = kept + incoming
                changed.append("allowances")
                stats["allowances"] += 1
        stats["files"] += 1
        if changed:
            snap["profile"] = profile
            path.write_text(
                json.dumps(snap, ensure_ascii=False, indent=2, default=_json_default) + "\n",
                encoding="utf-8",
            )
            stats["updated"] += 1
        touched.add(code)

    for path in emp_dir.glob("*.json"):
        code = path.stem
        if code in SKIP_TEST_CODES or code in touched:
            continue
        snap = json.loads(path.read_text(encoding="utf-8"))
        profile = snap.setdefault("profile", {})
        ident = apply_default_identity(profile)
        if ident:
            snap["profile"] = profile
            path.write_text(
                json.dumps(snap, ensure_ascii=False, indent=2, default=_json_default) + "\n",
                encoding="utf-8",
            )
            stats["identity_all"] += 1
    return stats


def write_extract(
    records: dict[str, dict[str, Any]],
    out_dir: Path,
    *,
    xlsx_path: Path,
    merge_stats: dict[str, int] | None = None,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    serializable = {code: rec for code, rec in sorted(records.items())}
    (out_dir / "employees.json").write_text(
        json.dumps(serializable, ensure_ascii=False, indent=2, default=_json_default) + "\n",
        encoding="utf-8",
    )

    import csv

    report_path = out_dir / "BAO_CAO.csv"
    with report_path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "msnv",
                "full_name",
                "gender",
                "marital_status",
                "children_count",
                "phone",
                "bank_account",
                "si_book_no",
                "id_number",
                "education_code",
            ]
        )
        for code, rec in sorted(serializable.items()):
            w.writerow(
                [
                    code,
                    rec.get("full_name") or "",
                    rec.get("gender") or "",
                    rec.get("marital_status") or "",
                    rec.get("children_count") if rec.get("children_count") is not None else "",
                    rec.get("phone") or "",
                    rec.get("bank_account") or "",
                    rec.get("si_book_no") or "",
                    rec.get("id_number") or "",
                    rec.get("education_code") or "",
                ]
            )

    married = sum(1 for r in records.values() if r.get("marital_status") == "married")
    single = sum(1 for r in records.values() if r.get("marital_status") == "single")
    with_child = sum(1 for r in records.values() if int(r.get("children_count") or 0) > 0)
    bank = sum(1 for r in records.values() if r.get("bank_account"))
    phone = sum(1 for r in records.values() if r.get("phone"))
    manifest = {
        "schema_version": 1,
        "source_file": xlsx_path.name,
        "extracted_at": datetime.now().isoformat(timespec="seconds"),
        "employee_count": len(records),
        "marital_married": married,
        "marital_single": single,
        "with_children": with_child,
        "bank_account": bank,
        "phone": phone,
        "merge_stats": merge_stats or {},
        "notes": [
            "Hôn nhân: 0 = single, -1 hoặc 1 = married (Excel True = -1).",
            "Không đè BHXH/STK đã có nếu Excel cắt số (float).",
            "Bỏ qua MSNV test 1604, 1718, 8851.",
        ],
    }
    (out_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
