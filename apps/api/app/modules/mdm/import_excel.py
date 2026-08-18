"""
Import Excel nhân sự — một bảng map cột (P3b / 06§6.6a).

Mẫu HR: GET /api/employees/import-template (đủ trường hồ sơ chính).
Cột cũ (16) vẫn nhận. team_code khuyến nghị luôn điền.
"""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.modules.mdm.models import Employee, LookupValue, Position
from app.modules.mdm.schemas import ImportResult
from app.modules.mdm.service import resolve_employee_team, seed_lookup_values

# Map header Excel → field nội bộ (CHỈ một bảng — không nhân đôi)
COLUMN_MAP: dict[str, str] = {
    "employee_code": "employee_code",
    "msnv": "employee_code",
    "mã nv": "employee_code",
    "ma nv": "employee_code",
    "full_name": "full_name",
    "họ tên": "full_name",
    "ho ten": "full_name",
    "họ và tên": "full_name",
    "gender": "gender",
    "giới tính": "gender",
    "gioi tinh": "gender",
    "id_number": "id_number",
    "cccd": "id_number",
    "số cccd": "id_number",
    "so cccd": "id_number",
    "bank_account": "bank_account",
    "stk": "bank_account",
    "tài khoản nh": "bank_account",
    "tai khoan nh": "bank_account",
    "pay_channel": "pay_channel",
    "kênh lương": "pay_channel",
    "kenh luong": "pay_channel",
    "team_code": "team_code",
    "mã tổ": "team_code",
    "ma to": "team_code",
    "department_code": "department_code",
    "mã bộ phận": "department_code",
    "ma bo phan": "department_code",
    "bo phan": "department_code",
    "position_title": "position_title",
    "chức vụ": "position_title",
    "chuc vu": "position_title",
    "position_code": "position_code",
    "mã chức vụ": "position_code",
    "join_date": "join_date",
    "ngày vào": "join_date",
    "ngay vao": "join_date",
    "resign_date": "resign_date",
    "ngày nghỉ": "resign_date",
    "ngay nghi": "resign_date",
    "ngày thôi việc": "resign_date",
    "contract_signed_at": "contract_signed_at",
    "ngày ký hđ": "contract_signed_at",
    "ngay ky hd": "contract_signed_at",
    "probation_salary": "probation_salary",
    "lương thử việc": "probation_salary",
    "luong thu viec": "probation_salary",
    "contract_salary": "contract_salary",
    "lương hđ": "contract_salary",
    "lương hợp đồng": "contract_salary",
    "luong hd": "contract_salary",
    "status": "status",
    "trạng thái": "status",
    "trang thai": "status",
    "phone": "phone",
    "sđt": "phone",
    "điện thoại": "phone",
    "dien thoai": "phone",
    "birth_date": "birth_date",
    "ngày sinh": "birth_date",
    "ngay sinh": "birth_date",
    "children_count": "children_count",
    "số con": "children_count",
    "so con": "children_count",
    "marital_status": "marital_status",
    "tình trạng hôn nhân": "marital_status",
    "hon nhan": "marital_status",
    "nationality_code": "nationality_code",
    "quốc tịch": "nationality_code",
    "quoc tich": "nationality_code",
    "ethnicity_code": "ethnicity_code",
    "dân tộc": "ethnicity_code",
    "dan toc": "ethnicity_code",
    "religion_code": "religion_code",
    "tôn giáo": "religion_code",
    "ton giao": "religion_code",
    "education_code": "education_code",
    "trình độ": "education_code",
    "trinh do": "education_code",
    "birth_place_code": "birth_place_code",
    "nơi sinh": "birth_place_code",
    "noi sinh": "birth_place_code",
    "id_issue_date": "id_issue_date",
    "ngày cấp cccd": "id_issue_date",
    "ngay cap cccd": "id_issue_date",
    "id_issue_place_code": "id_issue_place_code",
    "nơi cấp cccd": "id_issue_place_code",
    "noi cap cccd": "id_issue_place_code",
    "permanent_address": "permanent_address",
    "địa chỉ thường trú": "permanent_address",
    "dia chi thuong tru": "permanent_address",
    "temporary_address": "temporary_address",
    "địa chỉ tạm trú": "temporary_address",
    "dia chi tam tru": "temporary_address",
    "urgent_contact": "urgent_contact",
    "liên hệ khẩn": "urgent_contact",
    "lien he khan": "urgent_contact",
    "si_book_no": "si_book_no",
    "số sổ bhxh": "si_book_no",
    "so so bhxh": "si_book_no",
    "si_enrolled": "si_enrolled",
    "tham gia bhxh": "si_enrolled",
    "si_base_override": "si_base_override",
    "mức đóng bh": "si_base_override",
    "muc dong bh": "si_base_override",
    "union_fee_override": "union_fee_override",
    "phí công đoàn": "union_fee_override",
    "phi cong doan": "union_fee_override",
}

# Header tiếng Việt trên file mẫu — khớp COLUMN_MAP.
TEMPLATE_HEADERS: list[str] = [
    "MSNV",
    "Họ tên",
    "Mã tổ",
    "Mã bộ phận",
    "Chức vụ",
    "Giới tính",
    "Ngày sinh",
    "Điện thoại",
    "Tình trạng hôn nhân",
    "Số con",
    "Quốc tịch",
    "Dân tộc",
    "Tôn giáo",
    "Trình độ",
    "Nơi sinh",
    "Số CCCD",
    "Ngày cấp CCCD",
    "Nơi cấp CCCD",
    "Địa chỉ thường trú",
    "Địa chỉ tạm trú",
    "Liên hệ khẩn",
    "Ngày vào",
    "Ngày ký HĐ",
    "Lương HĐ",
    "Lương thử việc",
    "Kênh lương",
    "Trạng thái",
    "Tham gia BHXH",
    "Số sổ BHXH",
    "Tài khoản NH",
    "Mức đóng BH",
    "Phí công đoàn",
]

MAX_UPLOAD_BYTES = 10 * 1024 * 1024


def _fold(s: str) -> str:
    t = unicodedata.normalize("NFC", (s or "").strip()).lower()
    t = "".join(c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


_COLUMN_MAP_FOLD: dict[str, str] = {_fold(k): v for k, v in COLUMN_MAP.items()}


def _norm_header(val: Any) -> str:
    return unicodedata.normalize("NFC", str(val or "").strip().lower())


def _map_header(val: Any) -> str | None:
    n = _norm_header(val)
    if n in COLUMN_MAP:
        return COLUMN_MAP[n]
    folded = _fold(n)
    return _COLUMN_MAP_FOLD.get(folded) if folded else None


def _str(val: Any) -> str | None:
    if val is None or val == "":
        return None
    text = str(val).strip()
    return text or None


def _parse_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"ngày không hợp lệ: {val}")


def _parse_money(val: Any) -> Decimal:
    if val is None or val == "":
        return Decimal("0")
    if isinstance(val, Decimal):
        return val
    text = str(val).replace(",", "").replace(" ", "").strip()
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"tiền không hợp lệ: {val}") from exc


def _parse_money_opt(val: Any) -> Decimal | None:
    if val is None or val == "":
        return None
    n = _parse_money(val)
    return n if n > 0 else None


def _parse_int(val: Any, default: int = 0) -> int:
    if val is None or val == "":
        return default
    try:
        return int(Decimal(str(val).replace(",", "").strip()))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"số không hợp lệ: {val}") from exc


def _parse_gender(val: Any) -> str | None:
    text = _str(val)
    if not text:
        return None
    s = _fold(text)
    if s in ("f", "female", "nu", "nữ") or text.strip().lower() in ("nữ", "nu"):
        return "female"
    if s in ("m", "male", "nam"):
        return "male"
    return text


def _parse_pay_channel(val: Any) -> str:
    text = (_str(val) or "ATM").strip().upper()
    folded = _fold(val or "")
    if text in ("CASH", "TM") or "tien mat" in folded:
        return "CASH"
    if text in ("ATM", "CK") or "atm" in folded:
        return "ATM"
    return "ATM"


def _parse_status(val: Any) -> str:
    text = (_str(val) or "active").strip().lower()
    folded = _fold(val or "")
    aliases = {
        "active": "active",
        "chinh thuc": "active",
        "probation": "probation",
        "thu viec": "probation",
        "maternity": "maternity",
        "thai san": "maternity",
        "resigned": "resigned",
        "thoi viec": "resigned",
        "suspended": "suspended",
    }
    st = aliases.get(folded) or aliases.get(text) or text
    if st not in ("active", "probation", "resigned", "suspended", "maternity"):
        return "active"
    return st


def _parse_bool_vn(val: Any) -> bool | None:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return val
    folded = _fold(str(val))
    if folded in ("1", "true", "yes", "y", "x", "co", "có"):
        return True
    if folded in ("0", "false", "no", "n", "khong", "không"):
        return False
    return None


def _resolve_lookup_rows(rows: list[LookupValue], raw: Any) -> str | None:
    text = _str(raw)
    if not text or not rows:
        return None
    up = text.upper()
    for row in rows:
        if row.code.upper() == up:
            return row.code
    folded = _fold(text)
    for row in rows:
        if _fold(row.name) == folded:
            return row.code
    for row in rows:
        name_fold = _fold(row.name)
        if name_fold and len(name_fold) >= 3 and name_fold in folded:
            return row.code
    return None


def _resolve_lookup(db: Session, group: str, raw: Any) -> str | None:
    rows = db.query(LookupValue).filter(LookupValue.group_code == group).all()
    return _resolve_lookup_rows(rows, raw)


def _resolve_position(db: Session, raw: Any) -> tuple[str | None, str | None]:
    text = _str(raw)
    if not text:
        return None, None
    pos = db.get(Position, text)
    if pos is None:
        folded = _fold(text)
        for row in db.query(Position).all():
            if row.code.upper() == text.upper() or _fold(row.name) == folded:
                pos = row
                break
    if pos is None:
        return None, text
    return pos.code, pos.name


def build_employee_import_template() -> bytes:
    """File mẫu HR — header tiếng Việt, 1 dòng ví dụ, sheet hướng dẫn."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhap NV"
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(TEMPLATE_HEADERS)
    for col, _h in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.row_dimensions[1].height = 28
    ws.append(
        [
            "8810",
            "NGUYỄN VĂN MẪU",
            "T1",
            "SW1",
            "Công nhân",
            "Nam",
            "01/01/1998",
            "0901234567",
            "Độc thân",
            0,
            "Việt Nam",
            "Kinh",
            "Không",
            "Trung học phổ thông",
            "Tây Ninh",
            "079098001234",
            "15/06/2021",
            "Cục Cảnh sát QLHC về TTXH",
            "Ấp 1, xã A, Tây Ninh",
            "",
            "Nguyễn Thị B - 0900000000",
            "01/08/2026",
            "01/08/2026",
            6500000,
            5500000,
            "ATM",
            "Thử việc",
            "Có",
            "1234567890",
            "0123456789",
            "",
            "",
        ]
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_HEADERS))}2"
    gender_dv = DataValidation(type="list", formula1='"Nam,Nữ"', allow_blank=True)
    ws.add_data_validation(gender_dv)
    gender_dv.add("F2:F500")
    pay_dv = DataValidation(type="list", formula1='"ATM,Tiền mặt"', allow_blank=True)
    ws.add_data_validation(pay_dv)
    pay_dv.add("Z2:Z500")
    st_dv = DataValidation(
        type="list", formula1='"Chính thức,Thử việc,Thai sản"', allow_blank=True
    )
    ws.add_data_validation(st_dv)
    st_dv.add("AA2:AA500")
    si_dv = DataValidation(type="list", formula1='"Có,Không"', allow_blank=True)
    ws.add_data_validation(si_dv)
    si_dv.add("AB2:AB500")

    guide = wb.create_sheet("Huong dan")
    guide["A1"] = "Hướng dẫn nhập nhân viên hàng loạt"
    guide["A1"].font = Font(bold=True, size=14)
    lines = [
        "1. Giữ nguyên hàng 1 (tiêu đề). Không đổi tên cột.",
        "2. Cột bắt buộc: MSNV, Họ tên, Mã tổ.",
        "3. Ngày: dd/mm/yyyy (15/06/2021) hoặc yyyy-mm-dd.",
        "4. Giới tính: Nam / Nữ. Kênh lương: ATM / Tiền mặt.",
        "5. Trạng thái: Chính thức / Thử việc / Thai sản.",
        "6. Tham gia BHXH: Có / Không.",
        "7. Quốc tịch, dân tộc, tôn giáo, trình độ, nơi sinh, nơi cấp CCCD, hôn nhân: gõ tên như trên hồ sơ NV (dropdown portal).",
        "8. Lương: số VND, có thể có dấu phẩy.",
        "9. Xóa dòng ví dụ 8810 trước khi nạp file thật.",
        "10. File .xlsx tối đa 10MB. Nạp tại «Tạo nhân viên mới» hoặc danh sách NV → Nhập Excel.",
    ]
    for i, line in enumerate(lines, start=3):
        guide[f"A{i}"] = line
    guide.column_dimensions["A"].width = 110

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_employees_xlsx(db: Session, content: bytes, filename: str) -> ImportResult:
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Trợ Lý AI: file Excel tối đa 10MB.",
        )
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Trợ Lý AI: chỉ hỗ trợ file .xlsx.",
        )

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="Trợ Lý AI: không đọc được file Excel.",
        ) from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Trợ Lý AI: file Excel trống.") from None

    field_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _map_header(cell)
        if key and key not in field_index:
            field_index[key] = idx

    if "employee_code" not in field_index or "full_name" not in field_index:
        raise HTTPException(
            status_code=400,
            detail="Trợ Lý AI: thiếu cột bắt buộc employee_code (MSNV) và full_name.",
        )

    created = 0
    updated = 0
    errors: list[str] = []

    seed_lookup_values(db)
    lookup_by_group: dict[str, list[LookupValue]] = {}

    def lookup(group: str, raw: Any) -> str | None:
        if group not in lookup_by_group:
            lookup_by_group[group] = (
                db.query(LookupValue).filter(LookupValue.group_code == group).all()
            )
        return _resolve_lookup_rows(lookup_by_group[group], raw)

    for row_no, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(field: str) -> Any:
            i = field_index.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        try:
            code = str(cell("employee_code") or "").strip()
            name = str(cell("full_name") or "").strip()
            if not code or not name:
                raise ValueError("thiếu MSNV hoặc họ tên")

            team_code = cell("team_code")
            dept_code = cell("department_code")
            team = (
                resolve_employee_team(
                    db,
                    None,
                    str(team_code).strip() if team_code else None,
                    str(dept_code).strip() if dept_code else None,
                    required=False,
                )
                if team_code
                else None
            )

            values: dict[str, Any] = {
                "full_name": name,
                "gender": _parse_gender(cell("gender")),
                "id_number": _str(cell("id_number")),
                "bank_account": _str(cell("bank_account")),
                "pay_channel": _parse_pay_channel(cell("pay_channel")),
                "position_title": _str(cell("position_title")),
                "join_date": _parse_date(cell("join_date")),
                "contract_signed_at": _parse_date(cell("contract_signed_at")),
                "probation_salary": _parse_money(cell("probation_salary")),
                "contract_salary": _parse_money(cell("contract_salary")),
                "status": _parse_status(cell("status")),
                "phone": _str(cell("phone")),
            }
            if "resign_date" in field_index:
                values["resign_date"] = _parse_date(cell("resign_date"))
                if values["resign_date"] and values["status"] == "active":
                    values["status"] = "resigned"
            if team is not None:
                values["team_id"] = team.id

            pos_raw = cell("position_code") or cell("position_title")
            pos_code, pos_name = _resolve_position(db, pos_raw)
            if pos_code:
                values["position_code"] = pos_code
            if pos_name:
                values["position_title"] = pos_name

            extra: list[tuple[str, Any]] = [
                ("birth_date", _parse_date(cell("birth_date")) if "birth_date" in field_index else None),
                ("id_issue_date", _parse_date(cell("id_issue_date")) if "id_issue_date" in field_index else None),
                ("children_count", _parse_int(cell("children_count")) if "children_count" in field_index else None),
                ("permanent_address", _str(cell("permanent_address"))),
                ("temporary_address", _str(cell("temporary_address"))),
                ("urgent_contact", _str(cell("urgent_contact"))),
                ("si_book_no", _str(cell("si_book_no"))),
                (
                    "si_enrolled",
                    _parse_bool_vn(cell("si_enrolled")) if "si_enrolled" in field_index else None,
                ),
                (
                    "si_base_override",
                    _parse_money_opt(cell("si_base_override")) if "si_base_override" in field_index else None,
                ),
                (
                    "union_fee_override",
                    _parse_money_opt(cell("union_fee_override")) if "union_fee_override" in field_index else None,
                ),
                ("marital_status", lookup("marital_status", cell("marital_status"))),
                ("nationality_code", lookup("nationality", cell("nationality_code"))),
                ("ethnicity_code", lookup("ethnicity", cell("ethnicity_code"))),
                ("religion_code", lookup("religion", cell("religion_code"))),
                ("education_code", lookup("education_level", cell("education_code"))),
                ("birth_place_code", lookup("birth_place", cell("birth_place_code"))),
                ("id_issue_place_code", lookup("id_issue_place", cell("id_issue_place_code"))),
            ]
            for key, val in extra:
                if key not in field_index:
                    continue
                if val is not None:
                    values[key] = val
                elif key in ("si_enrolled",):
                    values[key] = False

            emp = db.query(Employee).filter(Employee.employee_code == code).first()
            if emp is None:
                emp = Employee(employee_code=code, **values)
                db.add(emp)
                created += 1
            else:
                for k, v in values.items():
                    setattr(emp, k, v)
                emp.deleted_at = None
                updated += 1
        except Exception as exc:  # noqa: BLE001 — gom lỗi từng dòng
            errors.append(f"Dòng {row_no}: {exc}")
            if len(errors) >= 50:
                errors.append("… dừng báo lỗi sau 50 dòng.")
                break

    db.commit()
    detail = (
        f"Trợ Lý AI: import xong — tạo {created}, cập nhật {updated}, "
        f"lỗi {len(errors)} dòng."
    )
    return ImportResult(created=created, updated=updated, errors=errors, detail=detail)
