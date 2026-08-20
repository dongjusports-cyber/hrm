"""Xuất Excel KPI Dongju Sports VN — ngày và tháng. Cột tự cấu hình (tham khảo file HQ)."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.modules.core.export_log import log_export
from app.modules.print.context import COMPANY
from app.modules.report.schemas import KpiTeamMonthRow
from app.modules.report.team_kpi import compute_day, compute_month, list_day_people, list_month_people

NAVY = PatternFill("solid", fgColor="1B4F8A")
PALE = PatternFill("solid", fgColor="D6EAF8")
OT = PatternFill("solid", fgColor="FDEBD0")
WEEKEND = PatternFill("solid", fgColor="EEF2F7")
WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
TITLE = Font(name="Calibri", size=14, bold=True, color="1B4F8A")
SUB = Font(name="Calibri", size=10, color="334155")
BODY = Font(name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin", color="94A3B8"),
    right=Side(style="thin", color="94A3B8"),
    top=Side(style="thin", color="94A3B8"),
    bottom=Side(style="thin", color="94A3B8"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _num(v: Decimal | int | float | None):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    return v


def _header(ws, title: str, subtitle: str, cols: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(cols, 4))
    ws.cell(1, 1, COMPANY["name_vi"]).font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(cols, 4))
    ws.cell(2, 1, title).font = Font(name="Calibri", size=12, bold=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(cols, 4))
    ws.cell(3, 1, subtitle).font = SUB
    return 5


def _paint_header_row(ws, row: int, labels: list[str]) -> None:
    for i, label in enumerate(labels, 1):
        cell = ws.cell(row, i, label)
        cell.fill = NAVY
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN


def _cell(ws, r, c, value, *, fill=None, num_format=None):
    cell = ws.cell(r, c, value)
    cell.font = BODY
    cell.border = THIN
    cell.alignment = Alignment(vertical="center")
    if fill is not None:
        cell.fill = fill
    if num_format:
        cell.number_format = num_format
    return cell


def _autosize(ws, min_w=9, max_w=28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col[:40]:
            if cell.value is None:
                continue
            width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _pct(v) -> float | None:
    n = _num(v)
    return None if n is None else n / 100.0


def build_day_xlsx(db: Session, work_date: date, *, user_id=None) -> tuple[bytes, str]:
    data = compute_day(db, work_date)
    people = list_day_people(db, work_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Theo to"
    dmy = f"{work_date.day:02d}/{work_date.month:02d}/{work_date.year}"
    labels = [
        "No",
        "Bộ phận",
        "Tổ",
        "Loại",
        "HC",
        "Có mặt",
        "Vắng",
        "Thiếu chấm",
        "Đi trễ",
        "Người OT",
        "Giờ OT",
        "OT sổ",
        "OT ngoài",
        "Giờ OT/người",
    ]
    start = _header(
        ws,
        f"KPI ngày {dmy} — chuyên cần & tăng ca theo tổ",
        "Nguồn vân tay DJ-HRM · OT = sổ + ngoài + CN + lễ · Không gõ tay",
        len(labels),
    )
    ws.cell(4, 1, f"Tổ OT {data.teams_with_ot} · Người OT {data.ot_people} · Giờ OT {float(data.ot_hours):g}").font = SUB
    _paint_header_row(ws, start, labels)
    r = start + 1
    for i, t in enumerate(data.teams, 1):
        fill = OT if t.ot_people > 0 else None
        vals = [
            i,
            t.department_name,
            t.team_name,
            t.category_label,
            t.headcount,
            t.present,
            t.absent,
            t.missing_punch,
            t.late_people,
            t.ot_people,
            _num(t.ot_hours),
            _num(t.ot_on_books_hours),
            _num(t.ot_external_hours),
            _num(t.ot_hours_per_person),
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws, r, c, v, fill=fill, num_format="0.00" if c >= 11 else None)
        r += 1
    tot_fill = PALE
    totals = [
        "",
        "Tổng CTY",
        "",
        "",
        data.headcount,
        data.present,
        data.absent,
        data.missing_punch,
        data.late_people,
        data.ot_people,
        _num(data.ot_hours),
        "",
        "",
        "",
    ]
    for c, v in enumerate(totals, 1):
        cell = _cell(ws, r, c, v, fill=tot_fill, num_format="0.00" if c == 11 else None)
        cell.font = Font(name="Calibri", size=10, bold=True)
    _autosize(ws)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(labels))}{r}"

    ws2 = wb.create_sheet("Nguoi OT")
    labels2 = [
        "MSNV",
        "Họ tên",
        "Bộ phận",
        "Tổ",
        "Có mặt",
        "Vào",
        "Ra",
        "Công (h)",
        "OT sổ",
        "OT ngoài",
        "OT tổng",
        "Trễ (phút)",
        "Mã nghỉ",
    ]
    _header(ws2, f"Người tăng ca {dmy}", "Chỉ NV có giờ OT > 0", len(labels2))
    _paint_header_row(ws2, 5, labels2)
    r = 6
    ot_people = [p for p in people if p.ot_hours > 0]
    for p in ot_people:
        vals = [
            p.employee_code,
            p.full_name,
            p.department_name,
            p.team_name,
            "Có" if p.present else "",
            p.first_in or "",
            p.last_out or "",
            _num(p.worked_hours),
            _num(p.ot_on_books_hours),
            _num(p.ot_external_hours),
            _num(p.ot_hours),
            p.late_minutes or None,
            p.leave_code or "",
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws2, r, c, v, num_format="0.00" if c in (8, 9, 10, 11) else None)
        r += 1
    _autosize(ws2)
    ws2.freeze_panes = "A6"

    ws3 = wb.create_sheet("Vang thieu cham")
    labels3 = ["MSNV", "Họ tên", "Bộ phận", "Tổ", "Thiếu chấm", "Mã nghỉ", "Có mặt"]
    _header(ws3, f"Vắng / thiếu chấm {dmy}", "Ngày công mà không có vân tay, hoặc có mã nghỉ", len(labels3))
    _paint_header_row(ws3, 5, labels3)
    r = 6
    for p in people:
        missing = (not p.present) or bool(p.leave_code)
        if not missing:
            continue
        vals = [
            p.employee_code,
            p.full_name,
            p.department_name,
            p.team_name,
            "Thiếu chấm" if not p.present and not p.leave_code else "",
            p.leave_code or "",
            "Có" if p.present else "",
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws3, r, c, v)
        r += 1
    _autosize(ws3)

    buf = BytesIO()
    wb.save(buf)
    filename = f"KPI_ngay_{work_date.isoformat()}.xlsx"
    if user_id is not None:
        log_export(
            db,
            user_id=user_id,
            kind="kpi_day",
            period=work_date.isoformat(),
            row_count=len(data.teams),
            filename=filename,
        )
    return buf.getvalue(), filename


def _month_day_headers(teams: list[KpiTeamMonthRow]) -> list[str]:
    if not teams:
        return []
    return [c.work_date[8:10] for c in teams[0].days]


def build_month_xlsx(db: Session, period: str, *, user_id=None) -> tuple[bytes, str]:
    data = compute_month(db, period)
    day_hdr = _month_day_headers(data.teams)

    wb = Workbook()
    ws = wb.active
    ws.title = "Chuyen can"
    labels = ["No", "Bộ phận", "Tổ", "Loại", "HC", *day_hdr, "Tổng ngày có mặt", "Nhân lực tháng", "Tỷ lệ chuyên cần"]
    start = _header(
        ws,
        f"KPI tháng {period} — chuyên cần theo tổ / ngày",
        data.formula_note,
        len(labels),
    )
    _paint_header_row(ws, start, labels)
    # weekend fill on header already navy; body uses WEEKEND
    weekend_att = set()
    weekend_ot = set()
    if data.teams:
        for i, cell in enumerate(data.teams[0].days):
            if not cell.is_workday:
                weekend_att.add(6 + i)
                weekend_ot.add(5 + i)
    r = start + 1
    for i, t in enumerate(data.teams, 1):
        vals = [i, t.department_name, t.team_name, t.category_label, t.headcount]
        for cell in t.days:
            vals.append(cell.present or None)
        vals.extend([_num(t.attendants), _num(t.monthly_manpower), _pct(t.attendance_rate_pct)])
        for c, v in enumerate(vals, 1):
            fill = WEEKEND if c in weekend_att else None
            fmt = "0.00%" if c == len(labels) else ("0.00" if c >= len(labels) - 2 else None)
            _cell(ws, r, c, v, fill=fill, num_format=fmt)
        r += 1
    _autosize(ws, min_w=5, max_w=22)
    ws.freeze_panes = "F6"
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(labels))}{max(r - 1, start)}"

    ws2 = wb.create_sheet("Tang ca")
    labels2 = [
        "No",
        "Bộ phận",
        "Tổ",
        "HC",
        *day_hdr,
        "Tổng giờ OT",
        "Người OT",
        "Giờ làm (có mặt×8h)",
        "Tỷ lệ OT (chia sẻ)",
        "Tỷ lệ OT (công suất)",
    ]
    _header(
        ws2,
        f"KPI tháng {period} — tăng ca theo tổ / ngày",
        "OT = sổ + ngoài + CN + lễ. Ô ngày = tổng giờ tổ.",
        len(labels2),
    )
    _paint_header_row(ws2, 5, labels2)
    r = 6
    for i, t in enumerate(data.teams, 1):
        vals = [i, t.department_name, t.team_name, t.headcount]
        for cell in t.days:
            vals.append(_num(cell.ot_hours) if cell.ot_hours else None)
        vals.extend(
            [
                _num(t.ot_hours),
                t.ot_people,
                _num(t.actual_work_hours),
                _pct(t.ot_share_pct),
                _pct(t.ot_capacity_pct),
            ]
        )
        for c, v in enumerate(vals, 1):
            fill = None
            if 5 <= c <= 4 + len(day_hdr) and isinstance(v, (int, float)) and v > 0:
                fill = OT
            if c in weekend_ot:
                fill = WEEKEND if fill is None else OT
            fmt = "0.00%" if c >= len(labels2) - 1 else ("0.00" if c >= 5 else None)
            _cell(ws2, r, c, v, fill=fill, num_format=fmt)
        r += 1
    _autosize(ws2, min_w=5, max_w=22)
    ws2.freeze_panes = "E6"

    ws3 = wb.create_sheet("Nghi viec")
    labels3 = ["No", "Bộ phận", "Tổ", "Đầu kỳ", "Tuyển", "Nghỉ", "Cuối kỳ", "Tỷ lệ nghỉ"]
    _header(ws3, f"KPI tháng {period} — biến động nhân sự theo tổ", "Begin / In / Out / End", len(labels3))
    _paint_header_row(ws3, 5, labels3)
    r = 6
    for i, t in enumerate(data.teams, 1):
        vals = [
            i,
            t.department_name,
            t.team_name,
            t.begin_hc,
            t.recruit,
            t.resign,
            t.end_hc,
            _pct(t.turnover_rate_pct),
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws3, r, c, v, num_format="0.00%" if c == 8 else None)
        r += 1
    for c, v in enumerate(
        ["", "Tổng CTY", "", data.begin_hc, data.recruit, data.resign, data.end_hc, _pct(data.turnover_rate_pct)],
        1,
    ):
        cell = _cell(ws3, r, c, v, fill=PALE, num_format="0.00%" if c == 8 else None)
        cell.font = Font(name="Calibri", size=10, bold=True)
    _autosize(ws3)

    ws4 = wb.create_sheet("Tong hop")
    _header(ws4, f"KPI tháng {period} — tổng hợp công ty", data.formula_note, 4)
    summary = [
        ("Kỳ", period),
        ("Ngày công chuẩn (B3)", _num(data.param_b3)),
        ("HC", data.headcount),
        ("Ngày có mặt (tổng)", _num(data.attendants)),
        ("Nhân lực tháng", _num(data.monthly_manpower)),
        ("Tỷ lệ chuyên cần", _pct(data.attendance_rate_pct)),
        ("Giờ OT", _num(data.ot_hours)),
        ("Người OT", data.ot_people),
        ("Giờ làm (có mặt×8h)", _num(data.actual_work_hours)),
        ("Tỷ lệ OT chia sẻ (file sếp)", _pct(data.ot_share_pct)),
        ("Tỷ lệ OT công suất", _pct(data.ot_capacity_pct)),
        ("Tuyển", data.recruit),
        ("Nghỉ việc", data.resign),
        ("Tỷ lệ nghỉ việc", _pct(data.turnover_rate_pct)),
    ]
    _paint_header_row(ws4, 5, ["Chỉ tiêu", "Giá trị"])
    r = 6
    for k, v in summary:
        _cell(ws4, r, 1, k, fill=PALE)
        fmt = "0.00%" if isinstance(k, str) and "Tỷ lệ" in k else ("0.00" if isinstance(v, float) else None)
        _cell(ws4, r, 2, v, num_format=fmt)
        r += 1
    _autosize(ws4, min_w=28, max_w=40)

    people_by_team = list_month_people(db, period)
    ws5 = wb.create_sheet("Nguoi OT thang")
    labels5 = ["MSNV", "Họ tên", "Bộ phận", "Tổ", "Ngày có mặt", "Ngày trễ", "Giờ OT"]
    _header(ws5, f"Người có OT tháng {period}", "Gộp cả tháng — lưu hồ sơ", len(labels5))
    _paint_header_row(ws5, 5, labels5)
    r = 6
    for p in people_by_team:
        if p.ot_hours <= 0:
            continue
        vals = [
            p.employee_code,
            p.full_name,
            p.department_name,
            p.team_name,
            p.present_days,
            p.late_days,
            _num(p.ot_hours),
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws5, r, c, v, num_format="0.00" if c == 7 else None)
        r += 1
    _autosize(ws5)

    buf = BytesIO()
    wb.save(buf)
    filename = f"KPI_thang_{period}.xlsx"
    if user_id is not None:
        log_export(
            db,
            user_id=user_id,
            kind="kpi_month",
            period=period,
            row_count=len(data.teams),
            filename=filename,
        )
    return buf.getvalue(), filename
