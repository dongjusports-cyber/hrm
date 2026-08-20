"""Xuất Excel bảng công tháng — GET chỉ đọc, không ensure kỳ / không rebuild."""

from __future__ import annotations

from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.modules.attendance.timesheet import list_timesheets
from app.modules.core.excel_filename import company_excel_filename


def export_timesheets_xlsx(
    db: Session,
    period: str,
    *,
    department_id: UUID | None = None,
    department_code: str | None = None,
    employee_code: str | None = None,
) -> tuple[bytes, str]:
    rows = list_timesheets(
        db,
        period,
        employee_code=employee_code,
        department_id=department_id,
        department_code=department_code,
    )
    extra = (employee_code or "").strip() or (department_code or "").strip().upper()
    if not extra and rows and rows[0].department_code and department_id is not None:
        extra = rows[0].department_code
    wb = Workbook()
    ws = wb.active
    ws.title = "Bang_cong"
    bold = Font(bold=True)
    scope = extra or "toàn công ty"
    ws.append([f"Bảng công kỳ {period} — {scope}"])
    ws.append([f"Số dòng: {len(rows)}"])
    if not rows:
        ws.append(["Chưa có kỳ lương hoặc chưa tổng hợp / không khớp lọc."])
    ws.append([])
    headers = [
        "STT",
        "MSNV",
        "Họ tên",
        "Bộ phận",
        "Công",
        "Phép năm",
        "Nghỉ REM",
        "Trễ",
        "Sớm",
        "Tăng ca sổ",
        "Tăng ca ngoài",
        "OT CN",
        "OT lễ",
    ]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
    for i, row in enumerate(rows, start=1):
        dept = " ".join(
            p for p in ((row.department_code or ""), (row.department_name or "")) if p
        ).strip()
        ws.append(
            [
                i,
                row.employee_code,
                row.full_name,
                dept,
                float(row.worked_days),
                float(row.al_days),
                float(row.rem_days),
                row.late_count,
                row.early_count,
                float(row.ot_hours_weekday),
                float(row.ot_hours_external),
                float(row.ot_hours_weekend),
                float(row.ot_hours_holiday),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), company_excel_filename("Bảng công", period=period, extra=extra or None)
