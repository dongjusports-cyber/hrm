"""Danh sách + Excel NV đã tích chu kỳ trong kỳ — không đổi công thức công."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.modules.attendance.engine import VN_TZ, to_vn
from app.modules.attendance.models import AttendanceDay
from app.modules.attendance.schemas import CycleLeaveRowOut
from app.modules.attendance.timesheet import get_pay_period, parse_period
from app.modules.mdm.models import Employee


def _fmt_hm(dt: datetime | None) -> str:
    if dt is None:
        return ""
    local = to_vn(dt).astimezone(VN_TZ)
    return local.strftime("%H:%M")


def list_cycle_leave(db: Session, period: str) -> list[CycleLeaveRowOut]:
    parse_period(period)
    pay = get_pay_period(db, period)
    if pay is None:
        return []
    rows = (
        db.query(AttendanceDay, Employee)
        .join(Employee, Employee.id == AttendanceDay.employee_id)
        .filter(
            AttendanceDay.cycle_leave.is_(True),
            AttendanceDay.work_date >= pay.date_from,
            AttendanceDay.work_date <= pay.date_to,
            Employee.deleted_at.is_(None),
        )
        .order_by(AttendanceDay.work_date, Employee.employee_code)
        .all()
    )
    return [
        CycleLeaveRowOut(
            employee_code=emp.employee_code,
            full_name=emp.full_name,
            work_date=day.work_date,
            first_in=day.first_in,
            last_out=day.last_out,
            worked_hours=day.worked_hours,
            note=day.note or "",
        )
        for day, emp in rows
    ]


def export_cycle_leave_xlsx(db: Session, period: str) -> tuple[bytes, str]:
    items = list_cycle_leave(db, period)
    wb = Workbook()
    ws = wb.active
    ws.title = "Chu_ky"
    bold = Font(bold=True)
    ws.append([f"Danh sách chu kỳ — kỳ {period}"])
    ws.append([f"Số lượt: {len(items)}"])
    ws.append([])
    headers = ["STT", "MSNV", "Họ tên", "Ngày về", "Vào", "Ra", "Giờ công", "Ghi chú"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
    for i, row in enumerate(items, start=1):
        ws.append(
            [
                i,
                row.employee_code,
                row.full_name,
                row.work_date.strftime("%d/%m/%Y"),
                _fmt_hm(row.first_in),
                _fmt_hm(row.last_out),
                float(row.worked_hours),
                row.note,
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"Chu_ky_{period}.xlsx"
