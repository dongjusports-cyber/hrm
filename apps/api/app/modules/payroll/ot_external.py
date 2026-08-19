"""
OT ngoài — tính tiền theo 22§22.8 (cùng ot_base, theo phút) nhưng KHÔNG vào payslip/BHXH/PIT.

Chi trả kênh ATM riêng (policy ot_split.exclude_from_payslip).
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.modules.attendance.models import PayPeriod, TimesheetMonth
from app.modules.attendance.timesheet import ensure_pay_period
from app.modules.core.export_log import log_export
from app.modules.core.models import User
from app.modules.mdm.models import Employee
from app.modules.payroll.engine_allowances import AllowanceInput, compute_allowances, should_zero_probation_allowances
from app.modules.payroll.engine_ot import (
    OtHours,
    OtInput,
    OtRateBuckets,
    buckets_from_parts,
    compute_ot_pay,
    hours_map_from_timesheet,
)
from app.modules.payroll.money import D, ZERO, money_vnd
from app.modules.payroll.period_eligibility import employee_on_payroll_period
from app.modules.print.context import COMPANY
from app.modules.payroll.service import (
    _active_policy,
    _detail_days_by_category,
    _leave_days_by_code,
    _monthly_map,
    _penalty_summary_for_employee,
    _type_views,
)


@dataclass(frozen=True)
class OtExternalPayRow:
    employee_code: str
    full_name: str
    bank_account: str
    raw_hours: Decimal
    effective_hours: Decimal
    ot_base: Decimal
    hourly_base: Decimal
    rate: Decimal
    amount_vnd: Decimal
    hours_x15: Decimal = ZERO
    pay_x15: Decimal = ZERO
    hours_x21: Decimal = ZERO
    pay_x21: Decimal = ZERO
    hours_x20: Decimal = ZERO
    pay_x20: Decimal = ZERO
    hours_x35: Decimal = ZERO
    pay_x35: Decimal = ZERO
    hours_x41: Decimal = ZERO
    pay_x41: Decimal = ZERO
    hours_x30: Decimal = ZERO
    pay_x30: Decimal = ZERO
    hours_x45: Decimal = ZERO
    pay_x45: Decimal = ZERO
    hours_x51: Decimal = ZERO
    pay_x51: Decimal = ZERO

    @property
    def buckets(self) -> OtRateBuckets:
        return OtRateBuckets(
            hours_x15=self.hours_x15,
            pay_x15=self.pay_x15,
            hours_x21=self.hours_x21,
            pay_x21=self.pay_x21,
            hours_x20=self.hours_x20,
            pay_x20=self.pay_x20,
            hours_x35=self.hours_x35,
            pay_x35=self.pay_x35,
            hours_x41=self.hours_x41,
            pay_x41=self.pay_x41,
            hours_x30=self.hours_x30,
            pay_x30=self.pay_x30,
            hours_x45=self.hours_x45,
            pay_x45=self.pay_x45,
            hours_x51=self.hours_x51,
            pay_x51=self.pay_x51,
        )


@dataclass(frozen=True)
class OtExternalSummary:
    period: str
    employee_count: int
    total_raw_hours: Decimal
    total_effective_hours: Decimal
    total_amount_vnd: Decimal
    rows: list[OtExternalPayRow]
    policy_note: str


_MONTH_EN = {
    1: "JANUARY",
    2: "FEBRUARY",
    3: "MARCH",
    4: "APRIL",
    5: "MAY",
    6: "JUNE",
    7: "JULY",
    8: "AUGUST",
    9: "SEPTEMBER",
    10: "OCTOBER",
    11: "NOVEMBER",
    12: "DECEMBER",
}

# In giống bảng lương: khối công ty + viền; màu theo yêu cầu in OT ngoài.
_RATE_PAIRS = [
    ("hours_x15", "pay_x15", "Hour x1.5", "Giờ x1.5\nT2–T7 17–22·6–8"),
    ("hours_x21", "pay_x21", "Hour x2.1", "Giờ x2.1\nT2–T7 22–6"),
    ("hours_x20", "pay_x20", "Hour x2", "Giờ x2\nCN 8–17"),
    ("hours_x35", "pay_x35", "Hour x3.5", "Giờ x3.5\nCN 17–22·6–8"),
    ("hours_x41", "pay_x41", "Hour x4.1", "Giờ x4.1\nCN 22–6"),
    ("hours_x30", "pay_x30", "Hour x3", "Giờ x3\nlễ 8–17"),
    ("hours_x45", "pay_x45", "Hour x4.5", "Giờ x4.5\nlễ 17–22·6–8"),
    ("hours_x51", "pay_x51", "Hour x5.1", "Giờ x5.1\nlễ 22–6"),
]
_LAST_COL = 7 + len(_RATE_PAIRS) * 2 + 3
_ROW_COMPANY = 1
_ROW_TITLE = 5
_ROW_PERIOD = 6
_ROW_NOTE = 7
_ROW_HDR_EN = 9
_ROW_HDR_VI = 10
_ROW_HDR_NUM = 11
_ROW_DATA = 12

_NAVY = "0A4D8C"
_LIGHT_BLUE = "BDD7EE"
_FOOTER_BLUE = "5B9BD5"
_WHITE = "FFFFFF"

_FILL_COMPANY = PatternFill(fill_type="solid", start_color=_NAVY, end_color=_NAVY)
_FILL_HEADER = PatternFill(fill_type="solid", start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE)
_FILL_NUM = PatternFill(fill_type="solid", start_color="D6EAF8", end_color="D6EAF8")
_FILL_FOOTER = PatternFill(fill_type="solid", start_color=_FOOTER_BLUE, end_color=_FOOTER_BLUE)

_FONT_COMPANY = Font(name="Arial", bold=True, size=14, color=_WHITE)
_FONT_META = Font(name="Arial", size=10, color="FFFFFF")
_FONT_TITLE = Font(name="Arial", bold=True, size=15, color=_NAVY)
_FONT_PERIOD = Font(name="Arial", bold=True, size=13, color=_NAVY)
_FONT_NOTE = Font(name="Arial", italic=True, size=9, color="5D6D7E")
_FONT_HDR = Font(name="Arial", bold=True, size=10, color="1A365D")
_FONT_NUM = Font(name="Arial", bold=True, size=9, color="1A365D")
_FONT_DATA = Font(name="Arial", size=10)
_FONT_FOOTER = Font(name="Arial", bold=True, size=10, color=_WHITE)

_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_R = Alignment(horizontal="right", vertical="center")

_THIN = Side(style="thin", color="7F8C8D")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADERS_EN = [
    "No",
    "Staff No",
    "Full Name",
    "Raw hours",
    "Hours (30 min)",
    "OT base",
    "Rate / hour",
]
_HEADERS_VI = [
    "STT",
    "MSNV",
    "Họ tên",
    "Giờ thô",
    "Giờ tính (30p)",
    "Nền OT",
    "Đơn giá/giờ",
]
for _h, _p, _en, _vi in _RATE_PAIRS:
    _HEADERS_EN.extend([_en, _en.replace("Hour", "Pay")])
    _HEADERS_VI.extend([_vi, _vi.replace("Giờ", "Tiền").split("\n")[0]])
_HEADERS_EN.extend(["OT pay (VND)", "Account No.", "Note"])
_HEADERS_VI.extend(["Tổng tiền OT", "Số tài khoản", "Ghi chú"])
_COL_WIDTHS = [6, 12, 26, 10, 12, 12, 12] + [11, 12] * len(_RATE_PAIRS) + [14, 16, 18]
_MONEY_COLS = frozenset({6, 7, _LAST_COL - 2} | {8 + i * 2 + 1 for i in range(len(_RATE_PAIRS))})
_HOUR_COLS = frozenset({4, 5} | {8 + i * 2 for i in range(len(_RATE_PAIRS))})
_CENTER_COLS = frozenset({1, 2})


def _period_title(period: str) -> str:
    year_s, month_s = period.split("-", 1)
    year, month = int(year_s), int(month_s)
    month_en = _MONTH_EN.get(month, f"M{month:02d}")
    return f"{month_en} {year} / THÁNG {month:02d} NĂM {year}"


def _style_header(cell, *, number_row: bool = False) -> None:
    cell.fill = _FILL_NUM if number_row else _FILL_HEADER
    cell.font = _FONT_NUM if number_row else _FONT_HDR
    cell.alignment = _ALIGN_C
    cell.border = _BORDER


def _style_data(cell, col: int) -> None:
    cell.font = _FONT_DATA
    cell.border = _BORDER
    if col in _CENTER_COLS:
        cell.alignment = _ALIGN_C
    elif col in _MONEY_COLS or col in _HOUR_COLS:
        cell.alignment = _ALIGN_R
    else:
        cell.alignment = _ALIGN_L


def _style_footer(cell) -> None:
    cell.font = _FONT_FOOTER
    cell.fill = _FILL_FOOTER
    cell.alignment = _ALIGN_C
    cell.border = _BORDER


def _external_rate(policy: dict) -> Decimal:
    ot_split = policy.get("ot_split") or {}
    ext = ot_split.get("ot_external") or {}
    rate_key = str(ext.get("rate_key", "weekday")).strip().lower()
    rates = policy.get("ot_rates") or {}
    return D(rates.get(rate_key, rates.get("weekday", "1.5")))


def split_external_ot_hours(
    *,
    external: Decimal,
    weekend: Decimal = ZERO,
    holiday: Decimal = ZERO,
) -> OtHours:
    """Tách giờ OT ngoài để tính tiền (ngày thường 1,5 · CN 2,0 · lễ 2–3).

    Sau rebuild: `external` = weekday_ext + CN + lễ.
    Dữ liệu cũ: `external` chỉ weekday_ext — nếu trừ CN/lễ ra âm thì giữ nguyên external.
    """
    ext = D(external)
    we = D(weekend)
    ho = D(holiday)
    wd = ext - we - ho
    if wd < 0:
        wd = ext
    return OtHours(weekday=wd, weekend=we, holiday=ho)


def _display_rate(hours: OtHours, policy: dict) -> Decimal:
    rates = policy.get("ot_rates") or {}
    pairs = [
        (hours.weekday, D(rates.get("weekday", "1.5"))),
        (hours.weekend, D(rates.get("weekend", rates.get("sunday", "2.0")))),
        (hours.holiday, D(rates.get("holiday", "2.0"))),
    ]
    best = max(pairs, key=lambda p: p[0])
    if best[0] <= 0:
        return _external_rate(policy)
    return best[1]


def _bucket_kwargs(b: OtRateBuckets) -> dict:
    return {
        "hours_x15": b.hours_x15,
        "pay_x15": b.pay_x15,
        "hours_x21": b.hours_x21,
        "pay_x21": b.pay_x21,
        "hours_x20": b.hours_x20,
        "pay_x20": b.pay_x20,
        "hours_x35": b.hours_x35,
        "pay_x35": b.pay_x35,
        "hours_x41": b.hours_x41,
        "pay_x41": b.pay_x41,
        "hours_x30": b.hours_x30,
        "pay_x30": b.pay_x30,
        "hours_x45": b.hours_x45,
        "pay_x45": b.pay_x45,
        "hours_x51": b.hours_x51,
        "pay_x51": b.pay_x51,
    }


def compute_ot_external_row(
    db: Session,
    pay: PayPeriod,
    emp: Employee,
    ts: TimesheetMonth,
    payload: dict,
) -> OtExternalPayRow | None:
    ext_map = hours_map_from_timesheet(ts, "external")
    if ext_map:
        hours = OtHours(by_rate=ext_map)
        raw = sum(ext_map.values(), ZERO)
    else:
        hours = split_external_ot_hours(
            external=D(ts.ot_hours_external),
            weekend=D(ts.ot_hours_weekend),
            holiday=D(ts.ot_hours_holiday),
        )
        raw = hours.weekday + hours.weekend + hours.holiday
    if raw <= 0:
        return None

    type_views = _type_views(db)
    monthly = _monthly_map(db, emp.id)
    penalty_sum = _penalty_summary_for_employee(db, pay, emp, payload)
    leave_days = _leave_days_by_code(db, pay.id, emp.id, ts.id)
    detail_days = _detail_days_by_category(db, ts.id)

    allow_res = compute_allowances(
        AllowanceInput(
            salary_divisor=pay.salary_divisor,
            worked_days=ts.worked_days,
            late_count=penalty_sum.late_count,
            early_count=penalty_sum.early_count,
            penalty_absent_days=penalty_sum.penalty_absent_days,
            join_date=emp.join_date,
            as_of=pay.date_to,
            policy=payload,
            monthly_by_code=monthly,
            types=type_views,
            leave_days_by_code=leave_days,
            detail_days_by_category=detail_days,
            penalty_audit=penalty_sum.detail,
            suppress_allowances=should_zero_probation_allowances(
                payload,
                contract_signed_at=emp.contract_signed_at,
                period_to=pay.date_to,
            ),
        )
    )

    ot_res = compute_ot_pay(
        OtInput(
            contract_salary=emp.contract_salary,
            salary_divisor=pay.salary_divisor,
            allowance_lines=allow_res.lines,
            attend_full_monthly=allow_res.attend_full_monthly,
            hours=hours,
            policy=payload,
        )
    )
    eff_map = ot_res.detail.get("effective_hours") or {}
    if ot_res.detail.get("time_bands"):
        eff = sum((D(v) for v in eff_map.values()), ZERO)
    else:
        eff = D(eff_map.get("weekday", "0")) + D(eff_map.get("weekend", "0")) + D(eff_map.get("holiday", "0"))
    buckets = buckets_from_parts(ot_res.detail.get("parts") or [])
    if eff <= 0 and ot_res.ot_pay <= 0:
        return OtExternalPayRow(
            employee_code=emp.employee_code,
            full_name=emp.full_name,
            bank_account=emp.bank_account or "",
            raw_hours=raw,
            effective_hours=ZERO,
            ot_base=ZERO,
            hourly_base=ZERO,
            rate=_display_rate(hours, payload) if not ext_map else D("1.5"),
            amount_vnd=ZERO,
            **_bucket_kwargs(buckets),
        )
    return OtExternalPayRow(
        employee_code=emp.employee_code,
        full_name=emp.full_name,
        bank_account=emp.bank_account or "",
        raw_hours=raw,
        effective_hours=eff,
        ot_base=ot_res.ot_base,
        hourly_base=ot_res.ot_hourly_base,
        rate=_display_rate(hours, payload) if not ext_map else D("1.5"),
        amount_vnd=ot_res.ot_pay,
        **_bucket_kwargs(buckets),
    )


def build_ot_external_summary(db: Session, period: str) -> OtExternalSummary:
    pay = ensure_pay_period(db, period)
    _, payload = _active_policy(db)
    ot_split = payload.get("ot_split") or {}
    ext_cfg = ot_split.get("ot_external") or {}
    note = str(
        ext_cfg.get(
            "note",
            "OT ngoài: x1,5 ngày thường · x2 CN/lễ≤8h · x2,1 đêm · x3 lễ>8h — không vào payslip/BHXH/PIT, chi ATM riêng.",
        )
    )

    q = (
        db.query(TimesheetMonth, Employee)
        .join(Employee, Employee.id == TimesheetMonth.employee_id)
        .filter(
            TimesheetMonth.pay_period_id == pay.id,
            or_(
                TimesheetMonth.ot_hours_external > 0,
                TimesheetMonth.ot_hours_weekend > 0,
                TimesheetMonth.ot_hours_holiday > 0,
            ),
            Employee.deleted_at.is_(None),
        )
        .order_by(Employee.employee_code.asc())
    )

    rows: list[OtExternalPayRow] = []
    for ts_row, emp in q.all():
        if not employee_on_payroll_period(emp, pay.date_from, pay.date_to):
            continue
        row = compute_ot_external_row(db, pay, emp, ts_row, payload)
        if row is not None:
            rows.append(row)

    total_raw = sum((r.raw_hours for r in rows), ZERO)
    total_eff = sum((r.effective_hours for r in rows), ZERO)
    total_amt = sum((r.amount_vnd for r in rows), ZERO)

    return OtExternalSummary(
        period=period,
        employee_count=len(rows),
        total_raw_hours=total_raw,
        total_effective_hours=total_eff,
        total_amount_vnd=money_vnd(total_amt),
        rows=rows,
        policy_note=note,
    )


def build_ot_external_excel(summary: OtExternalSummary) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "OT ngoài"

    def merge_row(row: int) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_LAST_COL)

    def paint_banner(row: int) -> None:
        for col in range(1, _LAST_COL + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = _FILL_COMPANY
            cell.alignment = _ALIGN_C
            cell.font = _FONT_META

    for r in (_ROW_COMPANY, _ROW_COMPANY + 1, _ROW_COMPANY + 2):
        paint_banner(r)
        merge_row(r)

    c_name = ws.cell(row=_ROW_COMPANY, column=1, value=COMPANY["name_vi"])
    c_name.font = _FONT_COMPANY
    c_name.fill = _FILL_COMPANY
    c_name.alignment = _ALIGN_C
    ws.row_dimensions[_ROW_COMPANY].height = 24

    c_addr = ws.cell(row=_ROW_COMPANY + 1, column=1, value=COMPANY["address_vi"])
    c_addr.font = _FONT_META
    c_addr.fill = _FILL_COMPANY
    c_addr.alignment = _ALIGN_C

    c_tel = ws.cell(row=_ROW_COMPANY + 2, column=1, value=f"Tel: {COMPANY['phone']}")
    c_tel.font = _FONT_META
    c_tel.fill = _FILL_COMPANY
    c_tel.alignment = _ALIGN_C

    merge_row(_ROW_TITLE)
    t1 = ws.cell(
        row=_ROW_TITLE,
        column=1,
        value="OT NGOÀI / BẢNG OT NGOÀI — CHI ATM RIÊNG",
    )
    t1.font = _FONT_TITLE
    t1.alignment = _ALIGN_C
    ws.row_dimensions[_ROW_TITLE].height = 22

    merge_row(_ROW_PERIOD)
    t2 = ws.cell(row=_ROW_PERIOD, column=1, value=_period_title(summary.period))
    t2.font = _FONT_PERIOD
    t2.alignment = _ALIGN_C
    ws.row_dimensions[_ROW_PERIOD].height = 20

    merge_row(_ROW_NOTE)
    note = ws.cell(
        row=_ROW_NOTE,
        column=1,
        value=summary.policy_note or "Không vào phiếu lương / BHXH / PIT — chi ATM riêng.",
    )
    note.font = _FONT_NOTE
    note.alignment = _ALIGN_C

    for col, val in enumerate(_HEADERS_EN, start=1):
        cell = ws.cell(row=_ROW_HDR_EN, column=col, value=val)
        _style_header(cell)
    for col, val in enumerate(_HEADERS_VI, start=1):
        cell = ws.cell(row=_ROW_HDR_VI, column=col, value=val)
        _style_header(cell)
    for col in range(1, _LAST_COL + 1):
        cell = ws.cell(row=_ROW_HDR_NUM, column=col, value=col)
        _style_header(cell, number_row=True)

    ws.row_dimensions[_ROW_HDR_EN].height = 22
    ws.row_dimensions[_ROW_HDR_VI].height = 28
    ws.row_dimensions[_ROW_HDR_NUM].height = 18

    hour_fmt = "0.00"
    money_fmt = "#,##0"

    totals: dict[str, Decimal] = {}
    for i, r in enumerate(summary.rows, start=1):
        excel_row = _ROW_DATA + i - 1
        pair_vals: list = []
        for h_attr, p_attr, _en, _vi in _RATE_PAIRS:
            hv = D(getattr(r, h_attr, 0))
            pv = D(getattr(r, p_attr, 0))
            totals[h_attr] = totals.get(h_attr, ZERO) + hv
            totals[p_attr] = totals.get(p_attr, ZERO) + pv
            pair_vals.extend([float(hv), int(pv)])
        values = [
            i,
            r.employee_code,
            r.full_name,
            float(r.raw_hours),
            float(r.effective_hours),
            int(r.ot_base),
            int(r.hourly_base),
            *pair_vals,
            int(r.amount_vnd),
            r.bank_account or "",
            f"OT ngoài {summary.period}",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            _style_data(cell, col)
            if col in _HOUR_COLS:
                cell.number_format = hour_fmt
            elif col in _MONEY_COLS:
                cell.number_format = money_fmt

    footer = _ROW_DATA + len(summary.rows)
    ws.merge_cells(start_row=footer, start_column=1, end_row=footer, end_column=3)
    fc = ws.cell(
        row=footer,
        column=1,
        value=f"Tổng cộng  ({summary.employee_count} nhân viên)",
    )
    _style_footer(fc)
    for col in range(2, 4):
        _style_footer(ws.cell(row=footer, column=col))

    footer_vals: dict[int, float | int] = {
        4: float(summary.total_raw_hours),
        5: float(summary.total_effective_hours),
        _LAST_COL - 2: int(summary.total_amount_vnd),
    }
    for idx, (h_attr, p_attr, _en, _vi) in enumerate(_RATE_PAIRS):
        hcol = 8 + idx * 2
        footer_vals[hcol] = float(totals.get(h_attr, ZERO))
        footer_vals[hcol + 1] = int(totals.get(p_attr, ZERO))
    for col in range(4, _LAST_COL + 1):
        cell = ws.cell(row=footer, column=col, value=footer_vals.get(col))
        _style_footer(cell)
        if col in _HOUR_COLS:
            cell.number_format = hour_fmt
            cell.alignment = _ALIGN_R
        elif col in _MONEY_COLS:
            cell.number_format = money_fmt
            cell.alignment = _ALIGN_R

    for idx, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = ws.cell(row=_ROW_DATA, column=1).coordinate
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9  # A4
    ws.print_title_rows = f"{_ROW_COMPANY}:{_ROW_HDR_NUM}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_ot_external(db: Session, period: str, actor: User) -> tuple[bytes, str]:
    summary = build_ot_external_summary(db, period)
    data = build_ot_external_excel(summary)
    filename = f"OT_ngoai_{period}.xlsx"
    log_export(
        db,
        user_id=actor.id,
        kind="ot_external",
        period=period,
        row_count=summary.employee_count,
        filename=filename,
    )
    db.commit()
    return data, filename
