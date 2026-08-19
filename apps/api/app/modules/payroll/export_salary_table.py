"""Xuất bảng lương theo mẫu GenusSuite (HIEN_PHAP/Salary/2.Salary table for *.xls).

Layout copy từ file Genus tháng 1–7/2026: 4 dòng header (EN + VI + số cột),
merge ô, nền xám header / bạc số cột, viền toàn bảng.
Số liệu lấy từ Payslip DJ-HRM — không từ GenusSuite.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.modules.attendance.models import PayPeriod, TimesheetMonth
from app.modules.attendance.timesheet import ensure_pay_period
from app.modules.mdm.models import Department, Employee, Team
from app.modules.payroll.engine_ot import OtRateBuckets
from app.modules.payroll.models import Payslip, PayslipComponent
from app.modules.payroll.money import D, ZERO, money_vnd
from app.modules.payroll.ot_external import compute_ot_external_row
from app.modules.payroll.period_eligibility import employee_on_payroll_period
from app.modules.payroll.service import _active_policy
from app.modules.print.context import COMPANY

MONTH_EN = {
    1: "JANUARY",
    2: "FEBRUARY",
    3: "MARCH",
    4: "APRIL",
    5: "MAY",
    6: "JUNE",
    7: "JULY",
    8: "AUGUST",
    9: "SEPTEMBER",
    10: "OCTOBER",
    11: "NOVEMBER",
    12: "DECEMBER",
}

# Header EN/VI/NUM — mẫu Genus + cột HSE tách khỏi PCCC (38 cột)
HEADER_EN_1 = [
    "No",
    "Section",
    "Staff No",
    "Full Name",
    "ID",
    "ACCOUNT\nNUMBER",
    "Sign",
    "Position",
    "D.O.J",
    "LC Sign date",
    "Probation\nSalary",
    "Confirmed\nSalary",
    "Working Day/ Ngày công",
    "",
    "",
    "Absent \nAL (Days)",
    "REM\n(Days)",
    "WD\nSalary",
    "ALLOWANCE / Phụ cấp",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "NOR OT / Tăng ca",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "GROSS\nSALARY",
    "S.I",
    "H.I",
    "U.I",
    "U.N",
    "Deduct ",
    "NET\nSALARY",
    "SEX",
]

HEADER_EN_2 = [
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "Pro\nThử việc",
    "OFF\nC. Thức",
    "Total\nT.Cộng",
    "",
    "",
    "",
    "Attend",
    "Position",
    "Toxic",
    "Trans",
    "PCCC",
    "HSE",
    "Tech",
    "Serverance",
    "Other",
    "Total",
    "Hour\nx1.5",
    "Pay\nx1.5",
    "Hour\nx2",
    "Pay\nx2",
    "Hour\nx2.1",
    "Pay\nx2.1",
    "Hour\nx3",
    "Pay\nx3",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
]

HEADER_ROW_VI = [
    "Stt",
    "Tổ",
    "MSNV",
    "Họ và Tên",
    "CMND",
    "Số TK",
    "Ký nhận",
    "Chức vụ",
    "Ngày vào",
    "Ngày ký",
    "Lương\nthử việc",
    "Lương\nhợp đồng",
    "",
    "",
    "",
    "Nghỉ Phép năm",
    "Nghỉ chế độ",
    "Lương ngày công",
    "Chuyên cần",
    "Chức vụ",
    "Độc hại",
    "Đi lại",
    "PCCC",
    "HSE",
    "Tay nghề may",
    "Thâm niên",
    "Khác",
    "Tổng",
    "Giờ x1.5\nngày thường",
    "Tiền x1.5",
    "Giờ x2\nCN · lễ ≤8h",
    "Tiền x2",
    "Giờ x2.1\nđêm",
    "Tiền x2.1",
    "Giờ x3\nlễ >8h",
    "Tiền x3",
    "Tổng\nthu nhập",
    "BHXH",
    "BHYT",
    "BHTN",
    "CD",
    "Khấu trừ ",
    "Thực lãnh",
    "Giới Tính",
]

# xlrd merged_cells (0-based, exclusive end) → header block Genus
HEADER_MERGES = [
    (9, 11, 0, 1),
    (9, 11, 1, 2),
    (9, 11, 2, 3),
    (9, 11, 3, 4),
    (9, 11, 4, 5),
    (9, 11, 5, 6),
    (9, 11, 6, 7),
    (9, 11, 7, 8),
    (9, 11, 8, 9),
    (9, 11, 9, 10),
    (9, 11, 10, 11),
    (9, 11, 11, 12),
    (9, 10, 12, 15),
    (9, 11, 15, 16),
    (9, 11, 16, 17),
    (9, 11, 17, 18),
    (9, 10, 18, 28),
    (9, 10, 28, 36),
    (9, 11, 36, 37),
    (9, 11, 37, 38),
    (9, 11, 38, 39),
    (9, 11, 39, 40),
    (9, 11, 40, 41),
    (9, 11, 41, 42),
    (9, 11, 42, 43),
    (9, 11, 43, 44),
    (10, 12, 12, 13),
    (10, 12, 13, 14),
    (10, 12, 14, 15),
]

ALLOW_COLS = {
    "ATTEND": 18,
    "POSITION": 19,
    "TOXIC": 20,
    "TRANSPORT": 21,
    "PCCC": 22,
    "HSE": 23,
    "TECH": 24,
    "SENIORITY": 25,
    "OTHER": 26,
}

ROW_COMPANY = 1
ROW_PAYROLL_TITLE = 7
ROW_PERIOD_TITLE = 8
ROW_HEADER_EN_1 = 10
ROW_HEADER_EN_2 = 11
ROW_HEADER_VI = 12
ROW_HEADER_NUM = 13
ROW_DATA_START = 14
LAST_COL = 44

NUMERIC_DATA_COLS = frozenset(range(1, 8)) | frozenset(range(10, 44))  # 1-based, trừ SEX

NAVY = "0A4D8C"
LIGHT_BLUE = "BDD7EE"
PALE_BLUE = "D6EAF8"
FOOTER_BLUE = "5B9BD5"
WHITE = "FFFFFF"

FILL_COMPANY = PatternFill(fill_type="solid", start_color=NAVY, end_color=NAVY)
FILL_HEADER = PatternFill(fill_type="solid", start_color=LIGHT_BLUE, end_color=LIGHT_BLUE)
FILL_NUM = PatternFill(fill_type="solid", start_color=PALE_BLUE, end_color=PALE_BLUE)
FILL_FOOTER = PatternFill(fill_type="solid", start_color=FOOTER_BLUE, end_color=FOOTER_BLUE)

FONT_COMPANY = Font(name="Arial", bold=True, size=14, color=WHITE)
FONT_COMPANY_META = Font(name="Arial", size=10, color=WHITE)
FONT_TITLE = Font(name="Arial", bold=True, size=15, color=NAVY)
FONT_HDR = Font(name="Arial", bold=True, size=10, color="1A365D")
FONT_NUM = Font(name="Arial", bold=True, size=9, color="1A365D")
FONT_FOOTER = Font(name="Arial", bold=True, size=10, color=WHITE)
FONT_DATA = Font(name="Arial", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = [
    5,
    14,
    8,
    26,
    14,
    14,
    6,
    12,
    11,
    11,
    11,
    11,
    8,
    8,
    8,
    8,
    8,
    12,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    10,
    12,
    10,
    10,
    10,
    8,
    10,
    12,
    8,
]


def _period_parts(period: str) -> tuple[int, int]:
    y, m = period.split("-")
    return int(y), int(m)


def _fmt_date(d: date | None) -> str:
    if d is None:
        return ""
    return d.strftime("%d/%m/%Y")


def _fmt_sex(gender: str | None) -> str:
    if not gender:
        return ""
    g = gender.strip().upper()
    if g == "M":
        return "Nam"
    if g == "F":
        return "Nữ"
    return gender


def _num(v: Decimal | int | float | None) -> float | int:
    if v is None:
        return 0
    n = float(D(v))
    if n == int(n):
        return int(n)
    return round(n, 2)


def _style_header_cell(cell, *, number_row: bool = False) -> None:
    cell.fill = FILL_NUM if number_row else FILL_HEADER
    cell.font = FONT_NUM if number_row else FONT_HDR
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_ALL


def _style_data_cell(cell, *, col: int) -> None:
    cell.font = FONT_DATA
    cell.border = BORDER_ALL
    if col in NUMERIC_DATA_COLS:
        cell.alignment = ALIGN_RIGHT if col >= 10 else ALIGN_CENTER
    else:
        cell.alignment = ALIGN_LEFT


def _style_footer_cell(cell) -> None:
    cell.font = FONT_FOOTER
    cell.fill = FILL_FOOTER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_ALL


def _merge_genus(ws, rlo: int, rhi: int, clo: int, chi: int) -> None:
    ws.merge_cells(
        start_row=rlo + 1,
        start_column=clo + 1,
        end_row=rhi,
        end_column=chi,
    )


def _paint_header_grid(ws) -> None:
    for row in (ROW_HEADER_EN_1, ROW_HEADER_EN_2, ROW_HEADER_VI):
        for col in range(1, LAST_COL + 1):
            _style_header_cell(ws.cell(row=row, column=col))
    for col in range(1, LAST_COL + 1):
        _style_header_cell(ws.cell(row=ROW_HEADER_NUM, column=col), number_row=True)


def _write_company_block(ws) -> None:
    for row in (ROW_COMPANY, ROW_COMPANY + 1, ROW_COMPANY + 2):
        for col in range(1, LAST_COL + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_COMPANY
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_COMPANY_META
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)

    c1 = ws.cell(row=ROW_COMPANY, column=1, value=COMPANY["name_vi"])
    c1.font = FONT_COMPANY
    c1.fill = FILL_COMPANY
    c1.alignment = ALIGN_CENTER
    ws.row_dimensions[ROW_COMPANY].height = 24

    c2 = ws.cell(row=ROW_COMPANY + 1, column=1, value=COMPANY["address_vi"])
    c2.font = FONT_COMPANY_META
    c2.fill = FILL_COMPANY
    c2.alignment = ALIGN_CENTER

    c3 = ws.cell(row=ROW_COMPANY + 2, column=1, value=f"Tel: {COMPANY['phone']}")
    c3.font = FONT_COMPANY_META
    c3.fill = FILL_COMPANY
    c3.alignment = ALIGN_CENTER


def _write_header_block(ws) -> None:
    for col, val in enumerate(HEADER_EN_1, start=1):
        ws.cell(row=ROW_HEADER_EN_1, column=col, value=val or None)
    for col, val in enumerate(HEADER_EN_2, start=1):
        ws.cell(row=ROW_HEADER_EN_2, column=col, value=val or None)
    for col, val in enumerate(HEADER_ROW_VI, start=1):
        ws.cell(row=ROW_HEADER_VI, column=col, value=val or None)
    for col in range(1, LAST_COL + 1):
        ws.cell(row=ROW_HEADER_NUM, column=col, value=col)

    for rlo, rhi, clo, chi in HEADER_MERGES:
        _merge_genus(ws, rlo, rhi, clo, chi)

    _paint_header_grid(ws)

    ws.row_dimensions[ROW_HEADER_EN_1].height = 25
    ws.row_dimensions[ROW_HEADER_EN_2].height = 22
    ws.row_dimensions[ROW_HEADER_VI].height = 32
    ws.row_dimensions[ROW_HEADER_NUM].height = 20


def _load_rows(
    db: Session,
    pay: PayPeriod,
    *,
    channel: str | None = None,
    department_id: UUID | None = None,
    employee_code: str | None = None,
) -> list[tuple[Payslip, Employee, TimesheetMonth, str]]:
    q = (
        db.query(Payslip, Employee, TimesheetMonth, Team)
        .join(Employee, Employee.id == Payslip.employee_id)
        .join(TimesheetMonth, TimesheetMonth.employee_id == Employee.id)
        .filter(
            Payslip.pay_period_id == pay.id,
            TimesheetMonth.pay_period_id == pay.id,
            Employee.deleted_at.is_(None),
        )
        .outerjoin(Team, Team.id == Employee.team_id)
    )
    if channel in ("ATM", "CASH"):
        q = q.filter(Employee.pay_channel == channel)
    if department_id is not None:
        q = q.filter(Employee.department_id == department_id)
    if employee_code:
        q = q.filter(Employee.employee_code == employee_code.strip())
    q = q.order_by(Employee.employee_code.asc())
    out: list[tuple[Payslip, Employee, TimesheetMonth, str]] = []
    for slip, emp, ts, team in q.all():
        if not employee_on_payroll_period(emp, pay.date_from, pay.date_to):
            continue
        team_name = team.name if team else ""
        out.append((slip, emp, ts, team_name))
    return out


def _allowances_by_code(db: Session, payslip_id: UUID) -> dict[str, Decimal]:
    rows = db.query(PayslipComponent).filter(PayslipComponent.payslip_id == payslip_id).all()
    out: dict[str, Decimal] = defaultdict(lambda: ZERO)
    for row in rows:
        if row.component_code in ALLOW_COLS:
            out[row.component_code] += D(row.amount)
    return dict(out)


def _wd_days(db: Session, payslip_id: UUID) -> tuple[Decimal, Decimal]:
    rows = (
        db.query(PayslipComponent)
        .filter(PayslipComponent.payslip_id == payslip_id, PayslipComponent.component_code == "WD")
        .all()
    )
    prob = ZERO
    off = ZERO
    for row in rows:
        qty = D(row.quantity or 0)
        if row.segment == "probation":
            prob += qty
        else:
            off += qty
    return prob, off


def _ot_trong_pay(db: Session, payslip_id: UUID, fallback: Decimal) -> Decimal:
    """Tiền OT trong phiếu (T3/T5 đến 20:00). CN/lễ không vào đây."""
    rows = (
        db.query(PayslipComponent)
        .filter(PayslipComponent.payslip_id == payslip_id, PayslipComponent.component_code == "OT")
        .all()
    )
    if not rows:
        return fallback
    trong = ZERO
    has_split = False
    for row in rows:
        note = (row.note or "").lower()
        if any(k in note for k in ("weekend", "sunday", "holiday", "night")):
            has_split = True
            continue
        if "weekday" in note:
            has_split = True
        trong += D(row.amount)
    if has_split:
        return money_vnd(trong)
    return fallback


def _ot_salary_buckets(
    db: Session,
    pay: PayPeriod,
    emp: Employee,
    ts: TimesheetMonth,
    slip: Payslip,
    payload: dict,
) -> OtRateBuckets:
    """Giờ + tiền 4 mốc: OT trong phiếu + OT ngoài (x2/x2.1/x3 chi ATM riêng, không cộng GROSS)."""
    trong = OtRateBuckets(
        hours_x15=D(ts.ot_hours_weekday),
        pay_x15=_ot_trong_pay(db, slip.id, D(slip.ot_pay)),
    )
    ngoai = compute_ot_external_row(db, pay, emp, ts, payload)
    if ngoai is None:
        return trong
    return trong.plus(ngoai.buckets)


def _build_data_row(
    db: Session,
    stt: int,
    slip: Payslip,
    emp: Employee,
    ts: TimesheetMonth,
    team_name: str,
    pay: PayPeriod,
    payload: dict,
) -> list:
    allow = _allowances_by_code(db, slip.id)
    prob_days, off_days = _wd_days(db, slip.id)
    total_days = D(ts.worked_days)
    if prob_days <= 0 and off_days <= 0:
        off_days = total_days

    row: list = [""] * LAST_COL
    row[0] = stt
    row[1] = team_name
    row[2] = int(emp.employee_code) if emp.employee_code.isdigit() else emp.employee_code
    row[3] = emp.full_name
    row[4] = emp.id_number or ""
    row[5] = emp.bank_account or emp.phone or ""
    row[6] = ""
    row[7] = emp.position_title or ""
    row[8] = _fmt_date(emp.join_date)
    row[9] = _fmt_date(emp.contract_signed_at)
    row[10] = _num(emp.probation_salary)
    row[11] = _num(emp.contract_salary)
    row[12] = _num(prob_days)
    row[13] = _num(off_days)
    row[14] = _num(total_days)
    row[15] = _num(ts.al_days)
    row[16] = _num(ts.rem_days)
    row[17] = _num(slip.wd_salary)
    for code, col in ALLOW_COLS.items():
        row[col] = _num(allow.get(code, ZERO))
    row[27] = _num(slip.allowance_total)
    ot = _ot_salary_buckets(db, pay, emp, ts, slip, payload)
    row[28] = _num(ot.hours_x15)
    row[29] = _num(ot.pay_x15)
    row[30] = _num(ot.hours_x20)
    row[31] = _num(ot.pay_x20)
    row[32] = _num(ot.hours_x21)
    row[33] = _num(ot.pay_x21)
    row[34] = _num(ot.hours_x30)
    row[35] = _num(ot.pay_x30)
    row[36] = _num(slip.gross)
    row[37] = _num(slip.bhxh)
    row[38] = _num(slip.bhyt)
    row[39] = _num(slip.bhtn)
    row[40] = _num(slip.union_fee)
    row[41] = _num(slip.other_deductions)
    row[42] = int(money_vnd(slip.net))
    row[43] = _fmt_sex(emp.gender)
    return row


def _write_sheet(
    ws,
    period: str,
    *,
    channel_label: str,
    data_rows: list[tuple[Payslip, Employee, TimesheetMonth, str]],
    db: Session,
    pay: PayPeriod,
    payload: dict,
) -> int:
    year, month = _period_parts(period)
    month_en = MONTH_EN.get(month, str(month).upper())
    title = f"{month_en} {year} / THÁNG {month:02d} NĂM {year}{channel_label}"

    _write_company_block(ws)

    ws.merge_cells(
        start_row=ROW_PAYROLL_TITLE,
        start_column=1,
        end_row=ROW_PAYROLL_TITLE,
        end_column=LAST_COL,
    )
    t1 = ws.cell(row=ROW_PAYROLL_TITLE, column=1, value="PAYROLL / BẢNG LƯƠNG")
    t1.font = FONT_TITLE
    t1.alignment = ALIGN_CENTER

    ws.merge_cells(
        start_row=ROW_PERIOD_TITLE,
        start_column=1,
        end_row=ROW_PERIOD_TITLE,
        end_column=LAST_COL,
    )
    t2 = ws.cell(row=ROW_PERIOD_TITLE, column=1, value=title)
    t2.font = FONT_TITLE
    t2.alignment = ALIGN_CENTER

    ws.row_dimensions[ROW_PAYROLL_TITLE].height = 22
    ws.row_dimensions[ROW_PERIOD_TITLE].height = 22

    _write_header_block(ws)

    total_net = 0
    for i, (slip, emp, ts, team_name) in enumerate(data_rows, start=1):
        row_vals = _build_data_row(db, i, slip, emp, ts, team_name, pay, payload)
        excel_row = ROW_DATA_START + i - 1
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            _style_data_cell(cell, col=col)
        total_net += int(money_vnd(slip.net))

    footer = ROW_DATA_START + len(data_rows)
    ws.merge_cells(start_row=footer, start_column=1, end_row=footer, end_column=10)
    fc = ws.cell(row=footer, column=1, value=f"Grant Total  ({len(data_rows)} Employees)")
    _style_footer_cell(fc)
    for col in range(2, 11):
        _style_footer_cell(ws.cell(row=footer, column=col))
    for col in range(11, 43):
        _style_footer_cell(ws.cell(row=footer, column=col))
    nc = ws.cell(row=footer, column=43, value=total_net)
    _style_footer_cell(nc)
    _style_footer_cell(ws.cell(row=footer, column=44))

    for idx, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = ws.cell(row=ROW_DATA_START, column=1).coordinate

    return len(data_rows)


def build_salary_table_xlsx(
    db: Session,
    period: str,
    channel: str = "ALL",
    *,
    department_id: UUID | None = None,
    employee_code: str | None = None,
) -> tuple[bytes, int, str]:
    """Trả (bytes, row_count, filename) — mẫu GenusSuite TOTAL + ATM + CASH."""
    pay = ensure_pay_period(db, period)
    ch = (channel or "ALL").upper()
    year, month = _period_parts(period)
    month_tag = MONTH_EN.get(month, f"M{month:02d}")
    code_filter = employee_code.strip() if employee_code else None
    if code_filter:
        probe = _load_rows(
            db, pay, department_id=department_id, employee_code=code_filter
        )
        if not probe:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Trợ Lý AI: không có phiếu lương MSNV {code_filter} kỳ {period}.",
            )

    wb = Workbook()
    first = wb.active
    assert first is not None
    row_count = 0
    _, payload = _active_policy(db)

    if ch == "ALL":
        first.title = "TOTAL"
        all_rows = _load_rows(
            db, pay, department_id=department_id, employee_code=code_filter
        )
        row_count += _write_sheet(
            first, period, channel_label="", data_rows=all_rows, db=db, pay=pay, payload=payload
        )

        ws_atm = wb.create_sheet("ATM")
        atm_rows = _load_rows(
            db, pay, channel="ATM", department_id=department_id, employee_code=code_filter
        )
        row_count += _write_sheet(
            ws_atm, period, channel_label=" (ATM)", data_rows=atm_rows, db=db, pay=pay, payload=payload
        )

        ws_cash = wb.create_sheet("CASH")
        cash_rows = _load_rows(
            db, pay, channel="CASH", department_id=department_id, employee_code=code_filter
        )
        row_count += _write_sheet(
            ws_cash, period, channel_label=" (CASH)", data_rows=cash_rows, db=db, pay=pay, payload=payload
        )
    elif ch == "ATM":
        first.title = "ATM"
        atm_rows = _load_rows(
            db, pay, channel="ATM", department_id=department_id, employee_code=code_filter
        )
        row_count = _write_sheet(
            first, period, channel_label=" (ATM)", data_rows=atm_rows, db=db, pay=pay, payload=payload
        )
    else:
        first.title = "CASH"
        cash_rows = _load_rows(
            db, pay, channel="CASH", department_id=department_id, employee_code=code_filter
        )
        row_count = _write_sheet(
            first, period, channel_label=" (CASH)", data_rows=cash_rows, db=db, pay=pay, payload=payload
        )

    scope_suffix = ""
    if code_filter:
        scope_suffix = f"_{code_filter}"
    elif department_id is not None:
        dept = db.get(Department, department_id)
        if dept:
            safe = dept.code or dept.name.replace(" ", "_")
            scope_suffix = f"_{safe}"

    filename = f"2.Salary table for {month_tag}.{year}{scope_suffix}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), row_count, filename
