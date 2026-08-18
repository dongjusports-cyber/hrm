"""
Import Excel phụ cấp — Go-live 10.9 / 03§3.4.
Hai dạng (tự nhận header):

1) Rộng (khuyến nghị):
   employee_code | ATTEND | TRANSPORT | TOXIC | POSITION | PCCC | TECH | SENIORITY | OTHER | CHILD

2) Dài:
   employee_code | allowance_code | amount

Ô trống = bỏ qua (không xóa gán cũ). Số 0 = gán 0.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator
from sqlalchemy.orm import Session

from app.modules.audit.service import write_audit
from app.modules.core.models import User
from app.modules.mdm.models import Employee
from app.modules.payroll.models import EmployeeAllowanceAssignment, PayComponent
from app.modules.payroll.money import money_vnd
from app.modules.payroll.seed_allowances import ASSIGNABLE_ALLOWANCE_CODES, CATALOG, seed_allowance_types

MAX_UPLOAD_BYTES = 10 * 1024 * 1024

CODE_ALIASES: dict[str, str] = {
    "employee_code": "_emp",
    "msnv": "_emp",
    "mã nv": "_emp",
    "ma nv": "_emp",
    "allowance_code": "_code",
    "mã phụ cấp": "_code",
    "ma phu cap": "_code",
    "code": "_code",
    "amount": "_amount",
    "số tiền": "_amount",
    "so tien": "_amount",
    # Wide columns
    "attend": "ATTEND",
    "chuyên cần": "ATTEND",
    "chuyen can": "ATTEND",
    "transport": "TRANSPORT",
    "đi lại": "TRANSPORT",
    "di lai": "TRANSPORT",
    "toxic": "TOXIC",
    "độc hại": "TOXIC",
    "doc hai": "TOXIC",
    "position": "POSITION",
    "chức vụ": "POSITION",
    "chuc vu": "POSITION",
    "pccc": "PCCC",
    "pccc+hse": "PCCC",
    "hse": "HSE",
    "tech": "TECH",
    "tay nghề": "TECH",
    "tay nghe": "TECH",
    "seniority": "SENIORITY",
    "thâm niên": "SENIORITY",
    "tham nien": "SENIORITY",
    "phone": "OTHER",
    "điện thoại": "OTHER",
    "dien thoai": "OTHER",
    "dt": "OTHER",
    "other": "OTHER",
    "khác": "OTHER",
    "khac": "OTHER",
    "child": "CHILD",
    "con nhỏ": "CHILD",
    "con nho": "CHILD",
}


class AllowanceImportResult(BaseModel):
    created: int
    updated: int
    skipped_empty: int
    errors: list[str]
    detail: str


class AllowanceAssignmentOut(BaseModel):
    id: UUID
    employee_code: str
    full_name: str
    allowance_code: str
    allowance_name: str
    amount: Decimal | None
    include_in_si_base: bool
    include_in_ot_base: bool


class AllowanceAssignmentUpsert(BaseModel):
    employee_code: str = Field(min_length=1, max_length=40)
    allowance_code: str = Field(min_length=1, max_length=40)
    amount: Decimal

    @field_validator("amount", mode="before")
    @classmethod
    def amount_dec(cls, v):  # noqa: ANN001
        if v is None or v == "":
            raise ValueError("Trợ Lý AI: cần số tiền phụ cấp.")
        raw = str(v).strip().replace(" ", "").replace(",", "")
        return Decimal(raw)


class AllowanceTypeOut(BaseModel):
    code: str
    name: str
    include_in_si_base: bool
    include_in_ot_base: bool
    default_amount: Decimal
    proration: str


def _norm(val: Any) -> str:
    return str(val or "").strip().lower()


def _parse_money(val: Any) -> Decimal | None:
    """None = ô trống (skip)."""
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return money_vnd(val)
    text = str(val).replace(",", "").replace(" ", "").strip()
    if text == "":
        return None
    try:
        return money_vnd(Decimal(text))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"tiền không hợp lệ: {val}") from exc


def _upsert(
    db: Session,
    emp: Employee,
    code: str,
    amount: Decimal,
    *,
    created: list[int],
    updated: list[int],
) -> None:
    at = db.query(PayComponent).filter(PayComponent.code == code).one_or_none()
    if at is None:
        raise ValueError(f"mã phụ cấp không có trong catalog: {code}")
    row = (
        db.query(EmployeeAllowanceAssignment)
        .filter(
            EmployeeAllowanceAssignment.employee_id == emp.id,
            EmployeeAllowanceAssignment.allowance_type_id == at.id,
        )
        .one_or_none()
    )
    if row is None:
        db.add(
            EmployeeAllowanceAssignment(
                employee_id=emp.id,
                allowance_type_id=at.id,
                amount=amount,
            )
        )
        db.flush()
        created[0] += 1
    else:
        row.amount = amount
        updated[0] += 1


def import_allowances_xlsx(
    db: Session,
    content: bytes,
    filename: str,
    *,
    actor: User | None = None,
) -> AllowanceImportResult:
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Trợ Lý AI: file Excel tối đa 10MB.",
        )
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Trợ Lý AI: chỉ hỗ trợ file .xlsx.",
        )

    seed_allowance_types(db)
    valid_codes = {c["code"] for c in CATALOG}

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="Trợ Lý AI: không đọc được file Excel.",
        ) from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Trợ Lý AI: file Excel trống.") from None

    mapped: list[str | None] = []
    for cell in header_row:
        mapped.append(CODE_ALIASES.get(_norm(cell)))

    if "_emp" not in mapped:
        raise HTTPException(
            status_code=400,
            detail="Trợ Lý AI: thiếu cột MSNV (employee_code).",
        )

    wide_cols = [c for c in mapped if c and c in valid_codes]
    long_mode = "_code" in mapped and "_amount" in mapped

    if not wide_cols and not long_mode:
        raise HTTPException(
            status_code=400,
            detail=(
                "Trợ Lý AI: cần cột phụ cấp (ATTEND/TRANSPORT/…) "
                "hoặc cặp allowance_code + amount."
            ),
        )

    created = [0]
    updated = [0]
    skipped = 0
    errors: list[str] = []
    emp_idx = mapped.index("_emp")

    for row_no, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            code = str(row[emp_idx] if emp_idx < len(row) else "").strip()
            if not code:
                raise ValueError("thiếu MSNV")
            emp = (
                db.query(Employee)
                .filter(Employee.employee_code == code, Employee.deleted_at.is_(None))
                .one_or_none()
            )
            if emp is None:
                raise ValueError(f"không tìm thấy MSNV {code}")

            if long_mode:
                c_idx = mapped.index("_code")
                a_idx = mapped.index("_amount")
                ac = str(row[c_idx] if c_idx < len(row) else "").strip().upper()
                amt = _parse_money(row[a_idx] if a_idx < len(row) else None)
                if amt is None:
                    skipped += 1
                else:
                    if ac not in valid_codes:
                        # alias single code
                        ac = CODE_ALIASES.get(ac.lower(), ac)
                    _upsert(db, emp, ac, amt, created=created, updated=updated)
            else:
                for col_i, field in enumerate(mapped):
                    if field not in valid_codes:
                        continue
                    if col_i >= len(row):
                        continue
                    amt = _parse_money(row[col_i])
                    if amt is None:
                        skipped += 1
                        continue
                    _upsert(db, emp, field, amt, created=created, updated=updated)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Dòng {row_no}: {exc}")
            if len(errors) >= 50:
                errors.append("… dừng báo lỗi sau 50 dòng.")
                break

    db.commit()
    detail = (
        f"Trợ Lý AI: import phụ cấp — tạo {created[0]}, cập nhật {updated[0]}, "
        f"bỏ trống {skipped}, lỗi {len(errors)} dòng."
    )
    if actor is not None:
        write_audit(
            db,
            actor=actor,
            action="payroll.allowances.import",
            entity_type="employee_allowance_assignments",
            entity_id=filename,
            summary=detail,
            meta={"created": created[0], "updated": updated[0], "errors": len(errors)},
        )
    return AllowanceImportResult(
        created=created[0],
        updated=updated[0],
        skipped_empty=skipped,
        errors=errors,
        detail=detail,
    )


def list_allowance_types(db: Session, *, assignable: bool = False) -> list[AllowanceTypeOut]:
    seed_allowance_types(db)
    query = db.query(PayComponent).filter(PayComponent.is_active.is_(True))
    if assignable:
        query = query.filter(PayComponent.code.in_(ASSIGNABLE_ALLOWANCE_CODES))
    rows = query.order_by(PayComponent.name.asc(), PayComponent.code.asc()).all()
    return [
        AllowanceTypeOut(
            code=r.code,
            name=r.name,
            include_in_si_base=r.include_in_si_base,
            include_in_ot_base=r.include_in_ot_base,
            default_amount=r.default_amount,
            proration=r.proration,
        )
        for r in rows
    ]


def list_assignments(
    db: Session,
    *,
    employee_code: str | None = None,
) -> list[AllowanceAssignmentOut]:
    seed_allowance_types(db)
    q = (
        db.query(EmployeeAllowanceAssignment, Employee, PayComponent)
        .join(Employee, Employee.id == EmployeeAllowanceAssignment.employee_id)
        .join(PayComponent, PayComponent.id == EmployeeAllowanceAssignment.allowance_type_id)
        .filter(Employee.deleted_at.is_(None))
        .order_by(Employee.employee_code.asc(), PayComponent.code.asc())
    )
    if employee_code:
        q = q.filter(Employee.employee_code == employee_code.strip())
    out: list[AllowanceAssignmentOut] = []
    for asg, emp, at in q.all():
        out.append(
            AllowanceAssignmentOut(
                id=asg.id,
                employee_code=emp.employee_code,
                full_name=emp.full_name,
                allowance_code=at.code,
                allowance_name=at.name,
                amount=asg.amount,
                include_in_si_base=at.include_in_si_base,
                include_in_ot_base=at.include_in_ot_base,
            )
        )
    return out


def _assignment_out(
    asg: EmployeeAllowanceAssignment, emp: Employee, at: PayComponent
) -> AllowanceAssignmentOut:
    return AllowanceAssignmentOut(
        id=asg.id,
        employee_code=emp.employee_code,
        full_name=emp.full_name,
        allowance_code=at.code,
        allowance_name=at.name,
        amount=asg.amount,
        include_in_si_base=at.include_in_si_base,
        include_in_ot_base=at.include_in_ot_base,
    )


def upsert_assignment(
    db: Session,
    body: AllowanceAssignmentUpsert,
    *,
    actor: User | None = None,
) -> AllowanceAssignmentOut:
    seed_allowance_types(db)
    emp = (
        db.query(Employee)
        .filter(
            Employee.employee_code == body.employee_code.strip(),
            Employee.deleted_at.is_(None),
        )
        .one_or_none()
    )
    if emp is None:
        raise HTTPException(
            status_code=404,
            detail=f"Trợ Lý AI: không tìm thấy MSNV {body.employee_code}.",
        )
    code = body.allowance_code.strip().upper()
    at = (
        db.query(PayComponent)
        .filter(PayComponent.code == code, PayComponent.is_active.is_(True))
        .one_or_none()
    )
    if at is None:
        raise HTTPException(
            status_code=404,
            detail=f"Trợ Lý AI: không tìm thấy phụ cấp '{code}'.",
        )
    amount = money_vnd(body.amount)
    row = (
        db.query(EmployeeAllowanceAssignment)
        .filter(
            EmployeeAllowanceAssignment.employee_id == emp.id,
            EmployeeAllowanceAssignment.allowance_type_id == at.id,
        )
        .one_or_none()
    )
    created = row is None
    if row is None:
        row = EmployeeAllowanceAssignment(
            employee_id=emp.id,
            allowance_type_id=at.id,
            amount=amount,
        )
        db.add(row)
    else:
        row.amount = amount
    db.commit()
    db.refresh(row)
    if actor:
        write_audit(
            db,
            actor=actor,
            action="allowance.assignment.upsert",
            entity_type="employee_allowance",
            entity_id=str(row.id),
            summary=f"{'Gán' if created else 'Sửa'} phụ cấp {code} cho {emp.employee_code}",
            meta={"employee_code": emp.employee_code, "allowance_code": code, "amount": str(amount)},
        )
    return _assignment_out(row, emp, at)


def delete_assignment(
    db: Session,
    assignment_id: UUID,
    *,
    actor: User | None = None,
) -> dict[str, str]:
    row = db.get(EmployeeAllowanceAssignment, assignment_id)
    if row is None:
        raise HTTPException(status_code=404, detail="Trợ Lý AI: không tìm thấy gán phụ cấp.")
    emp = db.get(Employee, row.employee_id)
    at = db.get(PayComponent, row.allowance_type_id)
    db.delete(row)
    db.commit()
    if actor:
        write_audit(
            db,
            actor=actor,
            action="allowance.assignment.delete",
            entity_type="employee_allowance",
            entity_id=str(assignment_id),
            summary=(
                f"Xóa phụ cấp {at.code if at else '?'} của "
                f"{emp.employee_code if emp else '?'}"
            ),
            meta={
                "employee_code": emp.employee_code if emp else None,
                "allowance_code": at.code if at else None,
            },
        )
    return {"detail": "Trợ Lý AI: đã xóa phụ cấp khỏi hồ sơ."}
