"""Xuất bảng lương ATM / CASH Excel (P5.2) — Decimal → đồng."""

from __future__ import annotations

from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.modules.attendance.timesheet import ensure_pay_period
from app.modules.core.export_log import log_export
from app.modules.core.models import User
from app.modules.mdm.models import Department, Employee
from app.modules.payroll.models import Payslip
from app.modules.payroll.money import money_vnd

CHANNELS = ("ATM", "CASH", "ALL")


def _rows_for_channel(
    db: Session, period: str, channel: str
) -> list[tuple[Payslip, Employee, str]]:
    pay = ensure_pay_period(db, period)
    q = (
        db.query(Payslip, Employee, Department)
        .join(Employee, Employee.id == Payslip.employee_id)
        .outerjoin(Department, Department.id == Employee.department_id)
        .filter(Payslip.pay_period_id == pay.id, Employee.deleted_at.is_(None))
        .order_by(Employee.employee_code.asc())
    )
    if channel in ("ATM", "CASH"):
        q = q.filter(Employee.pay_channel == channel)
    out: list[tuple[Payslip, Employee, str]] = []
    for slip, emp, dept in q.all():
        out.append((slip, emp, dept.name if dept else ""))
    return out


def _write_sheet(ws, period: str, channel: str, rows: list[tuple[Payslip, Employee, str]]) -> int:
    bold = Font(bold=True)
    if channel == "ATM":
        headers = [
            "STT",
            "MSNV",
            "Họ tên",
            "Bộ phận",
            "Số tài khoản",
            "Thực lãnh (VND)",
            "Kỳ",
            "Trạng thái phiếu",
        ]
    else:
        headers = [
            "STT",
            "MSNV",
            "Họ tên",
            "Bộ phận",
            "Thực lãnh (VND)",
            "Kỳ",
            "Trạng thái phiếu",
            "Ghi chú",
        ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold

    total = 0
    for i, (slip, emp, dept_name) in enumerate(rows, start=1):
        net = int(money_vnd(slip.net))
        total += net
        if channel == "ATM":
            ws.append(
                [
                    i,
                    emp.employee_code,
                    emp.full_name,
                    dept_name,
                    emp.bank_account or "",
                    net,
                    period,
                    slip.status,
                ]
            )
        else:
            note = "Không có số TK" if not emp.bank_account else ""
            ws.append(
                [
                    i,
                    emp.employee_code,
                    emp.full_name,
                    dept_name,
                    net,
                    period,
                    slip.status,
                    note,
                ]
            )
    ws.append([])
    if channel == "ATM":
        ws.append(["", "", "", "", "Tổng thực lãnh", total, period, ""])
    else:
        ws.append(["", "", "", "Tổng thực lãnh", total, period, "", ""])
    return len(rows)


def build_payroll_export_xlsx(
    db: Session, period: str, channel: str
) -> tuple[bytes, int, str]:
    ch = (channel or "ALL").upper()
    if ch not in CHANNELS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Trợ Lý AI: channel chỉ nhận ATM, CASH hoặc ALL.",
        )

    ensure_pay_period(db, period)
    wb = Workbook()
    first = wb.active
    assert first is not None
    row_count = 0

    if ch == "ATM":
        first.title = "ATM"
        row_count = _write_sheet(first, period, "ATM", _rows_for_channel(db, period, "ATM"))
    elif ch == "CASH":
        first.title = "CASH"
        row_count = _write_sheet(first, period, "CASH", _rows_for_channel(db, period, "CASH"))
    else:
        atm_rows = _rows_for_channel(db, period, "ATM")
        cash_rows = _rows_for_channel(db, period, "CASH")
        first.title = "Tong_hop"
        first.append(["Kỳ", period])
        first.append(["Số NV ATM", len(atm_rows)])
        first.append(["Số NV CASH", len(cash_rows)])
        first.append(["Tổng dòng xuất", len(atm_rows) + len(cash_rows)])
        first["A1"].font = Font(bold=True)
        ws_atm = wb.create_sheet("ATM")
        ws_cash = wb.create_sheet("CASH")
        row_count = _write_sheet(ws_atm, period, "ATM", atm_rows) + _write_sheet(
            ws_cash, period, "CASH", cash_rows
        )

    filename = f"luong_{period}_{ch.lower()}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), row_count, filename


def export_payroll_channel(
    db: Session, user: User, period: str, channel: str
) -> tuple[bytes, str]:
    data, row_count, filename = build_payroll_export_xlsx(db, period, channel)
    kind = f"payroll_{(channel or 'ALL').lower()}"
    log_export(
        db,
        user_id=user.id,
        kind=kind,
        period=period,
        row_count=row_count,
        filename=filename,
    )
    return data, filename
